#!/usr/bin/python3
# -*- coding: utf-8 -*-
import sys
from sys import platform as _platform
import os
import glob
import pathlib
from pathlib import Path
from datetime import *
from time import sleep, strftime, gmtime
import shutil

import logging
from PyQt5 import *
from PyQt5.QtCore import *
from PyQt5.QtCore import QObject, QIODevice, pyqtSignal
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtSerialPort import QSerialPort, QSerialPortInfo

import openpyxl

from main_app_ui import Ui_MainWindow as main_app_ui
from pymodbus.client.sync import ModbusSerialClient as ModbusClient

from dialog_setup_product import DialogSetupProduct
from dialog_num_input import DialogNumInput
from dialog_system_check import DialogSystemCheck

from socks2 import SocketMain

if _platform == "linux" or _platform == "linux2":
    import RPi.GPIO as GPIO
    from gpiozero import LED, Button

# --------------------------------------------------------------------------- #
# configure the client logging
# --------------------------------------------------------------------------- #
logging.basicConfig()
log = logging.getLogger('pymodbus.server')
log.setLevel(logging.ERROR)

# ----------------------------------------------------------
# GLOBALS
# ----------------------------------------------------------

# --------------------------------------------------------------------------- #
# specify slaves for query
# --------------------------------------------------------------------------- #
tmm = 0x01
tsg = 0x02
tmux_0 = 0x03
tmux_1 = 0x03
tmux_2 = 0x04
tmux_3 = 0x05
tcb_1 = 0x06
tcb_2 = 0x07
tcb_3 = 0x08
ndb_1 = 0x10
ndb_2 = 0x11
ndb_3 = 0x12
ndb_4 = 0x13

# --------------------------------------------------------------------------- #
# Buttons
# --------------------------------------------------------------------------- #
btn_start_input = 13
signal_interlock_input = 24
signal_interrupt_input = 6
sg_sync_signal_input = 21


# --------------------------------------------------------------------------- #
# configure the client logging
# --------------------------------------------------------------------------- #
logging.basicConfig()
log = logging.getLogger('pymodbus.server')
log.setLevel(logging.ERROR)


class MainAppGui(QMainWindow):
    sg_sync_signal_detected = pyqtSignal(int)
    start_event_detected = pyqtSignal(int)
    interrupt_event_detected = pyqtSignal(int)
    interlock_event_detected = pyqtSignal(int)

    self_show_event_request = pyqtSignal(int)
    self_hide_event_request = pyqtSignal(int)

    def __init__(self, parent=None):
        QMainWindow.__init__(self, parent)
        self.ui = main_app_ui()
        self.ui.setupUi(self)
        self.setWindowTitle("MAIN APP")
        self.window_width = 1280
        self.window_height = 1024
        self.setFixedSize(self.window_width, self.window_height)
        self.move(0, 0)
        self.setWindowFlags(Qt.FramelessWindowHint)
        #self.setWindowFlags(Qt.Window | Qt.WindowTitleHint | Qt.CustomizeWindowHint )
        #self.showFullScreen()

        # QApplication.setOverrideCursor(Qt.BlankCursor)
        self.self_show_event_request.connect(self.show_main_window_event_handler)
        self.self_hide_event_request.connect(self.hide_main_window_event_handler)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.timer_event)

        self.app_ver = ""               # will change after loading config file
        self.serial_com_port_name = "COM2"  # will change after loading config file
        self.baudrate = 38400           # will change after loading config file
        self.mbTimeout = 0.5            # will change after loading config file
        self.mbRetries = 10             # will change after loading config file
        self.sg_ip_adr = "x.x.x.x"      # will change after loading config file
        self.sg_com_port = "COMx"       # will change after loading config file
        self.sg_connection = "xxx"      # will change after loading config file
        self.mm_ip_adr = "x.x.x.x"      # will change after loading config file
        self.mm_com_port = "COMx"       # will change after loading config file
        self.mm_connection = "xxx"      # will change after loading config file
        self.tester_id = ""             # will change after loading config file
        self.main_log_file_path = ""    # will change after loading config file
        self.debug_mode = False

        self.load_config_file()         # Load configuration from config.txt

        self.is_lid_closed = False
        self.is_lid_open = True
        self.is_start_button_pressed = False
        self.is_start_button_released = True

        # RPI IO and Buttons setup
        if _platform == "linux" or _platform == "linux2":
            print("mount NAS folder Raspberry pi")
            os.system("sudo mount -t cifs -o user=tester_00,password=StateCuba,sec=ntlm,rw,vers=1.0,file_mode=0777,dir_mode=0777 //192.168.201.81/tester_00/log /home/pi/wa/GenericTester_excell/log/nas")

            GPIO.setmode(GPIO.BCM)  # Set board mode to Broadcom

            GPIO.setup(btn_start_input, GPIO.IN, pull_up_down=GPIO.PUD_UP)  # Set pin 13: Start button
            self.start_event_detected.connect(self.btn_start_click)
            GPIO.add_event_detect(btn_start_input, GPIO.FALLING, callback=self.start_event_detected.emit, bouncetime=100)

            GPIO.setup(sg_sync_signal_input, GPIO.IN)  # Set pin 21: SG sync out
            self.sg_sync_signal_detected.connect(self.sg_sync_signal_event_handler)
            GPIO.add_event_detect(sg_sync_signal_input, GPIO.RISING, callback=self.sg_sync_signal_detected.emit, bouncetime=100)

            GPIO.setup(signal_interlock_input, GPIO.IN, pull_up_down=GPIO.PUD_UP)  # Set pin 24: Interlock switch
            self.interlock_event_detected.connect(self.interlock_event_handler)
            # GPIO.add_event_detect(signal_interlock, GPIO.BOTH, callback=self.interlock_event_detected.emit, bouncetime=100)

            self.is_lid_closed = (GPIO.input(signal_interlock_input) == 0)
            self.is_lid_open = (GPIO.input(signal_interlock_input) == 1)


        dlg = DialogSystemCheck()
        dlg.setWindowTitle("System check and initialisation")

        dlg.show()
        dlg.ui.lbl_lbl_progress.repaint()
        dlg.ui.pbProgress.repaint()
        dlg.ui.lbl_check_info.repaint()
        dlg.ui.lbl_lbl_mm_port.repaint()
        dlg.ui.lbl_mm_port.repaint()
        dlg.ui.lbl_lbl_sg_port.repaint()
        dlg.ui.lbl_sg_port.repaint()
        dlg.repaint()

        # init modbus communication
        self.modbus_client = ModbusClient(method='ascii', port=self.serial_com_port_name, baudrate=self.baudrate,
                                          timeout=self.mbTimeout, retries=self.mbRetries)
        dlg.ui.lbl_check_info.setText("Modbus initialisation on port " + self.serial_com_port_name)
        dlg.ui.lbl_check_info.repaint()

        try:
            self.modbus_client.connect()
        except Exception as e:
            log.debug(e)
            log.debug("\r\nError initialising modbus library")
            # self.close()
            # sys.exit(app.exec_())
        self.previous_dut_button_object = ""

        if self.sg_connection == "ser" or self.mm_connection == "ser":
            self.port_list = QSerialPortInfo.availablePorts()
            dlg.setWindowTitle("Detecting instruments")

        self.mm_id_cmd = "*idn?\n"
        # Multimeter communication
        if self.mm_connection == "ser":
            self.mm_ser = QSerialPort()
            self.mm_ser.setBaudRate(9600)
            mm_found = False
            portName = self.mm_com_port
            self.mm_ser.setPortName(portName)
            try:
                self.mm_ser.open(QIODevice.ReadWrite)
                if self.mm_ser.isOpen():
                    sleep(1)
                    self.mm_ser.flush()
                    mm_id = str(self.mm_read(self.mm_id_cmd, True))
                    if mm_id.find("1908") != -1:
                        print("MM found on port: ", portName)
                        mm_found = True
                        dlg.ui.lbl_mm_port.setText(portName)
                        dlg.ui.lbl_mm_port.repaint()
                        print("MM info: %s" % mm_id)
            except Exception as e:
                pass
            if not mm_found:
                for port in self.port_list:
                    try:
                        portName = port.portName()
                        self.mm_ser.setPortName(portName)
                        self.mm_ser.open(QIODevice.ReadWrite)
                        if self.mm_ser.isOpen():
                            print(portName)
                            sleep(1)
                            self.mm_ser.flush()
                            mm_id = str(self.mm_read(self.mm_id_cmd, True))
                            if mm_id.find("1908") != -1:
                                print("MM found on port: ", portName)
                                mm_found = True
                                dlg.ui.lbl_mm_port.setText(portName)
                                dlg.ui.lbl_mm_port.repaint()
                                print("MM info: %s" % mm_id)
                                self.port_list.remove(port)
                                break
                            else:
                                self.mm_ser.close()
                    except Exception as e:
                        pass
            if not mm_found:
                QMessageBox.critical(self, "Warning", "No multimeter found on system")
                self.close()
                exit(1)
        elif self.mm_connection == "tcp":
            print("MM is on TCP...")
        else:
            QMessageBox.critical(self, "Warning", "No multimeter setup on system.\n Check config.txt file.")
            self.close()
            exit(1)

        self.sg_id_cmd = "*idn?\n"
        # Signal generator configuration
        if self.sg_connection == "ser":
            self.sg_ser = QSerialPort()
            self.sg_ser.setBaudRate(115200)
            sg_found = False
            portName = self.sg_com_port
            self.sg_ser.setPortName(portName)
            try:
                self.sg_ser.setPortName(portName)
                self.sg_ser.open(QIODevice.ReadWrite)
                if self.sg_ser.isOpen():
                    sleep(1)
                    sg_id = str(self.sg_send(self.sg_id_cmd, True))
                    if sg_id.find("TG2511A") != -1:
                        print("SG founf on port: ", portName)
                        sg_found = True
                        print("SG info: %s" % sg_id)
                    else:
                        self.sg_ser.close()
            except Exception as e:
                pass
            if not sg_found:
                print("2nd attempt, trying to find SG on all ports ...", end="")
                for port in self.port_list:
                    print(portName)
                    try:
                        portName = port.portName()
                        self.sg_ser.setPortName(portName)
                        self.sg_ser.open(QIODevice.ReadWrite)
                        if self.sg_ser.isOpen():
                            sleep(1)
                            sg_id = str(self.sg_send(self.sg_id_cmd, True))
                            if sg_id.find("TG2511A") != -1:
                                print("SG founf on port: ", portName)
                                sg_found = True
                                print("SG info: %s" % sg_id)
                                dlg.ui.lbl_sg_port.setText(portName)
                                dlg.ui.lbl_sg_port.repaint()
                                break
                            else:
                                self.sg_ser.close()
                    except Exception as e:
                        pass
            if not sg_found:
                QMessageBox.critical(self, "Warning", "No No signal generator found on system")
                self.close()
                exit(1)
        elif self.sg_connection == "tcp":
            print("Signal  generator is on TCP...")
        else:
            dlg.close()
            QMessageBox.critical(self, "Warning", "No signal generator setup on system.\n Check config.txt file.")
            self.close()
            exit(1)
        dlg.close()

        # Application system variables
        self.tmm_outputs = [False] * 3
        self.tmux_1_outputs = [False] * 8
        self.tmux_2_outputs = [False] * 8
        self.tmux_3_outputs = [False] * 8
        self.tcb_1_outputs = [False] * 16
        self.tcb_2_outputs = [False] * 16
        self.tcb_3_outputs = [False] * 16
        self.ndb_1_outputs = [False] * 24
        self.ndb_2_outputs = [False] * 24
        self.ndb_3_outputs = [False] * 24
        self.ndb_4_outputs = [False] * 24

        self.sg_cmd = ""
        self.mm_last_cmd = ""
        self.mm_value = ""
        self.last_value = 0
        self.mm_units = ""
        self.pMax = ""
        self.pMin = ""
        self.mm_last_range = 0
        self.is_mm_result = False
        self.is_mm_error = False
        self.is_enabled_log_mm_result = False
        self.sg_commands = []
        self.test_lines = []
        self.product_config_lines = []
        self.test_number = 0
        self.units_tested = 0
        self.comm_counter = 0

        # Logging data

        self.log_lines = []
        self.mm_log_lines = []
        self.tt_log_lines = []
        self.barcode_slot_1 = self.barcode_slot_2 = self.barcode_slot_3 = self.barcode_slot_4 = ""
        self.is_valid_board_on_slot_1 = self.is_valid_board_on_slot_2 = self.is_valid_board_on_slot_3 = self.is_valid_board_on_slot_4 = False
        self.is_bar_code_on_slot = [False] * 4
        self.is_board_on_slot = [False] * 4
        self.barcode = "00000000000000000"
        self.product_name = ""
        self.nest_id = ""
        self.insert_id = ""
        self.grn_no = "Verification"
        self.insert_number_duts = 0
        self.insert_template_name = ""
        self.user_id = ""
        self.slot = 1
        self.freq = 50
        self.fc_setup = ""
        self.pass_min = ""
        self.pass_max = ""
        self.is_trip_error = False
        self.is_trip_result = False
        self.TripTime = 0
        self.TripType = ""
        self.duration = ""
        self.sg_setup_num = ""
        self.sp1 = ""
        self.sp2 = ""
        self.sp3 = ""
        self.sp4 = ""
        self.sp5 = ""
        self.sp6 = ""
        self.sp7 = ""
        self.test_pass_result_slot_1 = [True] * 30
        self.test_pass_result_slot_2 = [True] * 30
        self.test_pass_result_slot_3 = [True] * 30
        self.test_pass_result_slot_4 = [True] * 30
        self.test_value_result_slot_1 = [0] * 30
        self.test_value_result_slot_2 = [0] * 30
        self.test_value_result_slot_3 = [0] * 30
        self.test_value_result_slot_4 = [0] * 30
        self.is_slot_1_test_pass = False
        self.is_slot_2_test_pass = False
        self.is_slot_3_test_pass = False
        self.is_slot_4_test_pass = False
        self.fail_units_counters_per_test = [0] * 30
        self.fail_units_counters_per_test_for_slot_1 = [0] * 30
        self.fail_units_counters_per_test_for_slot_2 = [0] * 30
        self.fail_units_counters_per_test_for_slot_3 = [0] * 30
        self.fail_units_counters_per_test_for_slot_4 = [0] * 30

        self.pass_units_counter_global = 0
        self.fail_units_counter_global = 0
        self.loopCounter = 0

        # QMessageBox.information(self, '', self.get_num())
        self.ui.sbLoopCount.setEnabled(False)

        # Prevent second test precudure execution
        self.test_in_progress = False
        self.is_soft_button = False


        # self.timer.start(5000)
        # self.btn_start_click()

        self.data_folder = ""
        self.test_sequence_file = ""
        self.test_sequence_file_issue = ""
        self.test_sequence_file_ref = ""
        self.test_specification_file = ""
        self.test_specification_file_issue = ""
        self.test_specification_file_ref = ""
        self.test_calibration_file = ""
        self.test_calibration_file_ref = ""
        self.test_sequence_lines = []
        self.test_specification_lines = []
        self.product_calibration_lines = []
        self.exec_test = []
        self.tests = [[], []]

        self.btn_back_click()

        self.debug_log_file_name = ""
        self.file_name_new_format = ""
        self.log_file_name_new_format = ""
        self.log_nas_file_name = ""
        self.comm_error_log_file_name = ""
        self.bc_error_log_file_name = ""
        self.mm_error_log_file_name = ""

        # # import template for middle part of screen.
        # # This part of screen show test sequence progress
        # ui = eval("t_" + self.product_name + ".Ui_Frame()")
        # ui.setupUi(self.ui.template_frame)

        self.create_new_log_files()

    def timer_event(self):
        self.timer.stop()
        self.btn_soft_start_click()

    def closeEvent(self):
        self.modbus_client.close()
        if self.sg_connection == "ser":
            self.sg_ser.close()
        else:
            pass
        if self.mm_connection == "ser":
            self.mm_ser.close()
        else:
            pass
        print("See you :)")
        self.close()
        sys.exit(app.exec_())

    def change_color(self, ui_item):
        template_css = """QPushButton { border-style: ridge;  border: 2px solid DarkSlateGray;  border-radius: 6px;  background-color: rgb(236, 236, 236);}"""
        command = "self.ui." + self.previous_dut_button_object + ".setStyleSheet('" + template_css + "')"
        exec(command)
        template_css = """QPushButton { border-style: ridge; border: 2px solid DarkSlateGray; border-radius: 6px; background-color: LightGreen }"""
        command = "self.ui." + ui_item + ".setStyleSheet('" + template_css + "')"
        exec(command)
        self.previous_dut_button_object = ui_item

    def set_btn_on_off_color(self, objekat, state):
        template_css = ""
        if state:
            template_css += """QPushButton { border-style: ridge; border: 2px solid DarkSlateGray; border-radius: 6px; background-color: LightGreen }"""
        else:
            template_css += """QPushButton { border-style: ridge;  border: 2px solid DarkSlateGray; border-radius: 6px; background-color: rgb(236, 236, 236);}"""
        command = "self.ui." + objekat + ".setStyleSheet('" + template_css + "')"
        exec(command)

    def set_btn_pass_fail_color(self, objekat, state):
        template_css = ""
        if state:
            template_css += """QPushButton { border-style: ridge; border: 2px solid DarkSlateGray; border-radius: 6px; background-color: LightGreen;}"""
        else:
            template_css += """QPushButton { border-style: ridge;  border: 2px solid DarkSlateGray; border-radius: 6px; background-color: LightSalmon;}"""
        command = "self.ui." + objekat + ".setStyleSheet('" + template_css + "')"
        exec(command)

    def set_slot_pass_fail_color(self, objekat, color):
        template_css = ""
        if color == 0:
            template_css += """"""
        if color == 1:
            template_css += """background-color: LightGreen"""
        if color == 2:
            template_css += """background-color: Crimson"""
        if color == 3:
            template_css += """background-color: LightSalmon"""
        if color == 4:
            template_css += """background-color: LightSkyBlue"""
        command = "self.ui." + objekat + ".setStyleSheet('" + template_css + "')"
        exec(command)

    def set_btn_color(self, item, color):
        template_css = ""
        if color == 0:
            template_css += """QPushButton { border-style: ridge;  border: 2px solid DarkSlateGray; border-radius: 6px; background-color: rgb(236, 236, 236);}"""
        if color == 1:
            template_css += """QPushButton { border-style: ridge; border: 2px solid DarkSlateGray; border-radius: 6px; background-color: LightGreen }"""
        if color == 2:
            template_css += """QPushButton { border-style: ridge; border: 2px solid DarkSlateGray; border-radius: 6px; background-color: Crimson; color:White }"""
        if color == 3:
            template_css += """QPushButton { border-style: ridge; border: 2px solid DarkSlateGray; border-radius: 6px; background-color: LightSalmon }"""
        if color == 4:
            template_css += """QPushButton { border-style: ridge; border: 2px solid DarkSlateGray; border-radius: 6px; background-color: LightSkyBlue }"""
        if color == 5:
            template_css += """QPushButton { border-style: ridge; border: 2px solid DarkSlateGray; border-radius: 6px; background-color: NavajoWhite }"""

        command = "self.ui." + item + ".setStyleSheet('" + template_css + "')"
        exec(command)
        command = "self.ui." + item + ".repaint()"
        exec(command)

    def load_config_file(self):
        # print("Load config file")
        file_name = "config.txt"
        file = open(file_name, "r")
        file_lines = [line.rstrip('\n') for line in file]
        file.close()
        for line in file_lines:
            line = line.replace(" ", "")
            # line = line.lower()
            if line[0] != "#" or line != "":
                parameters = line.split('=')

                if parameters[0] == "app_ver":
                    self.app_ver = parameters[1] == "1"

                if parameters[0] == "debug_mode":
                    self.debug_mode = parameters[1] == "1"

                if parameters[0] == "485port":
                    # """ Com port """
                    self.serial_com_port_name = parameters[1]
                    # print("\tCom_port_name = ", self.serial_com_port_name)

                if parameters[0] == "BaudRate":
                    """ Baudrate """
                    self.baudrate = parameters[1]
                    # print("\tBaudrate = ", self.baudrate)

                if parameters[0] == "timeout_s":
                    """ Modbus Timeout """
                    # self.mbTimeout = str(line[line.find("=")+1:])
                    exec("self.mbTimeout  = " + parameters[1])
                    # print("\tModbus Timeout = ", self.mbTimeout)

                if parameters[0] == "retries":
                    """ Modbus retries """
                    # self.mbTimeout = str(line[line.find("=")+1:])
                    exec("self.mbRetries  = " + parameters[1])
                    # print("\tModbus Timeout = ", self.mbTimeout)

                if parameters[0] == "sg_ip_adr":
                    """ Signal generator TCP-IP address """
                    self.sg_ip_adr = parameters[1]
                    # print("\tSignal generator TCP-IP address = ", self.sg_ip_adr)

                if parameters[0] == "sg_con":
                    """ Multimeter generator type  TCP or SER"""
                    self.sg_connection = parameters[1].lower()
                    # print("\tSignal generator connection type = ", self.sg_connection)

                if parameters[0] == "mm_ip_adr":
                    """ Multimeter TCP-IP address """
                    self.mm_ip_adr = parameters[1]
                    # print("\tMultimeter TCP-IP address = ", self.mm_ip_adr)

                if parameters[0] == "mm_com_port":
                    """ Multimeter generator com port """
                    self.mm_com_port = parameters[1]
                    # print("\tMultimeter com port = ", self.mm_com_port)

                if parameters[0] == "sg_com_port":
                    """ Signal generator com port """
                    self.sg_com_port = parameters[1]
                    # print("\tSignal generator com port = ", self.sg_com_port)

                if parameters[0] == "mm_con":
                    """ Multimeter generator type  TCP or SER"""
                    self.mm_connection = parameters[1].lower()
                    # print("\tMultimeter connection type = ", self.mm_connection)

                if parameters[0] == "tester_id":
                    self.tester_id = parameters[1]


    @staticmethod
    def get_num():
        dlg = DialogNumInput("text=55")
        dlg.setWindowTitle("Enter trip time in mS")
        dlg.exec_()
        text = dlg.text
        dlg.destroy()
        return text

    def btn_back_click(self):
        if _platform == "linux" or _platform == "linux2":
            GPIO.remove_event_detect(signal_interlock_input)
        dlg = DialogSetupProduct(self.nest_id)
        dlg.setWindowTitle("Product setup")
        dlg.exec_()
        action = dlg.action
        if action == 2:
            if _platform == "linux" or _platform == "linux2":
                os.system("sudo shutdown -h now")
        elif action == 3:
            self.closeEvent()
        else:
            self.user_id = dlg.user_name
            self.product_name = dlg.product_name
            self.nest_id = dlg.nest_id
            self.insert_id = dlg.insert_id
            self.grn_no = dlg.grn_no
            self.insert_number_duts = int(dlg.insert_number_duts)
            if self.insert_number_duts < 2:
                self.hide_dut(2)
            if self.insert_number_duts < 3:
                self.hide_dut(3)
            if self.insert_number_duts < 4:
                self.hide_dut(4)
            self.ui.lbl_product.setText(self.product_name)
            self.ui.lbl_insert.setText(self.insert_id)
            self.ui.lbl_grn.setText(self.grn_no)
            self.data_folder = ("products/" + self.product_name + "/")
            self.test_sequence_file = self.data_folder + self.product_name +" Test Sequence.xlsx"
            self.load_test_sequence_file()
            self.load_test_specification_file()
            self.load_test_calibration_file()

        if _platform == "linux" or _platform == "linux2":
            # Restore event detection for start signal_interlock
            GPIO.add_event_detect(signal_interlock_input, GPIO.BOTH, callback=self.interlock_event_detected.emit, bouncetime=100)

    def twInfoBoard_click(self):
        print("twInfoBoard_click")

    def create_new_log_files(self):
        # Create new LOG file
        dt = str(datetime.now())
        dt = dt[dt.find(" "):]
        dt = dt[:dt.find(".")]
        dt = dt[:len(dt) - 3]
        self.ui.lbl_date_time.setText(dt)
        dt = str(datetime.now())
        dt = dt[:dt.find(" ")]
        # Create new fornat CSV file
        self.log_file_name_new_format = "log/" + self.product_name + "_" + dt + "_" + self.tester_id + ".csv"
        # self.log_file_name_new_format = self.main_log_file_path + self.product_name + "_" + dt + "_" + self.tester_id + ".csv"
        self.file_name_new_format = self.product_name + "_" + dt + "_" + self.tester_id + ".csv"
        self.log_file_name_new_format = "log/" + self.file_name_new_format
        if not (os.path.isfile(self.log_file_name_new_format) and os.access(self.log_file_name_new_format, os.R_OK)):
            file = open(self.log_file_name_new_format, "w+")
            hdr_line = "Date & Time,Tester ID,Operator,Product,Nest ID,Insert ID,GRN,Spec_issue,Spec_Ref,Seq_issue,Seq_ref,Cal_ref,App ver,Slot,Bar Code," + \
                       "Test 1,Test 2,Test 3,Test 4,Test 5,Test 6,Test 7,Test 8,Test 9,Test 10," + \
                       "Test 11,Test 12,Test 13,Test 14,Test 15,Test 16,Test 17,Test 18,Test 19,Test 20," + \
                       "Test 21,Test 22,Test 23,Test 24,Test 25,Test 26,Test 27,Test 28,Test 29,Test 30," + \
                       "Failed tests\n"
            file.write(hdr_line)
            file.close()

        # Create fulle debug LOG file
        self.debug_log_file_name = "log/debug/debug_" + self.product_name + "_" + dt + "_" + self.tester_id + ".log"
        if not (os.path.isfile(self.debug_log_file_name) and os.access(self.debug_log_file_name, os.R_OK)):
            # create file
            file = open(self.debug_log_file_name, "w+")
            file.close()

        # Create new Communication error LOG file
        self.comm_error_log_file_name = "log/debug/comm_error_" + self.product_name + "_" + dt + "_" + self.tester_id + ".csv"
        if not (os.path.isfile(self.comm_error_log_file_name) and os.access(self.comm_error_log_file_name, os.R_OK)):
            file = open(self.comm_error_log_file_name, "w+")
            file.close()

        # Create new Bar code error LOG file
        self.bc_error_log_file_name = "log/debug/bc_error_" + self.product_name + "_" + dt + "_" + self.tester_id + ".csv"
        if not (os.path.isfile(self.bc_error_log_file_name) and os.access(self.bc_error_log_file_name, os.R_OK)):
            file = open(self.bc_error_log_file_name, "w+")
            file.close()

        # Create new Multimeter error LOG file
        self.mm_error_log_file_name = "log/debug/mm_error_" + self.product_name + "_" + dt + "_" + self.tester_id + ".csv"
        if not (os.path.isfile(self.mm_error_log_file_name) and os.access(self.mm_error_log_file_name, os.R_OK)):
            file = open(self.mm_error_log_file_name, "w+")
            
    # Backup log file will copy log file name t 
    def backup_log_file(self):
        self.log_nas_file_name = "log/nas/" + self.file_name_new_format
        shutil.copyfile(self.log_file_name_new_format, self.log_nas_file_name)

    def repaint_gui(self):
        # print("Cell and screen clean refresh...", end="")
        for test_slot in range(1, self.insert_number_duts + 1):
            # Clear red/green rectangles (pass/fail)
            self.set_slot_pass_fail_color("fSlot_" + str(test_slot), 0)
            self.set_pf_leds(False, False, test_slot)
            exec("self.ui.lbl_lbl_test_no_" + str(test_slot) + ".setVisible(True)")
            exec("self.ui.lbl_test_no_" + str(test_slot) + ".setVisible(True)")
            exec("self.ui.lbl_lbl_measurement_" + str(test_slot) + ".setVisible(True)")
            exec("self.ui.lbl_measurement_" + str(test_slot) + ".setVisible(True)")

            exec("self.ui.lbl_test_no_" + str(test_slot) + ".setText('?')")
            exec("self.ui.lbl_measurement_" + str(test_slot) + ".setText('?')")
            exec("self.ui.lbl_ResultTest_" + str(test_slot) + ".setText('?')")

            for test_num in range(1, 50):
                try:
                    pb_test = "pb_" + str(test_slot) + "_" + str(test_num)
                    exec_str = 'self.set_btn_on_off_color("' + pb_test + '", False)'
                    exec(exec_str)
                    exec("self.ui." + pb_test + ".repaint()")
                except Exception as e:
                    pass
        # print("done.")
        
    def hide_dut(self, slot_num):
        print("Hiding dut " + str(slot_num), end="")
        # Clear red/green rectangles (pass/fail)
        exec("self.ui.fSlot_" + str(slot_num) + ".setVisible(False)")
        exec("self.ui.lbl_lbl_product_" + str(slot_num) + ".setVisible(False)")
        exec("self.ui.lbl_bar_code_" + str(slot_num) + ".setVisible(False)")
        # exec("self.ui.lbl_lbl_test_no_" + str(slot_num) + ".setVisible(False)")
        # exec("self.ui.lbl_test_no_" + str(slot_num) + ".setVisible(False)")
        # exec("self.ui.lbl_lbl_measurement_" + str(slot_num) + ".setVisible(False)")
        # exec("self.ui.lbl_measurement_" + str(slot_num) + ".setVisible(False)")
        # exec("self.ui.lbl_test_no_" + str(slot_num) + ".setVisible(False)")
        # exec("self.ui.lbl_measurement_" + str(slot_num) + ".setVisible(False)")
        # exec("self.ui.lbl_ResultTest_" + str(slot_num) + ".setVisible(False)")
        for test_num in range(1, 31):
            try:
                exec_str = "self.ui.pb_" + str(slot_num) + "_" + str(test_num) + ".setVisible(False)"
                exec(exec_str)
            except Exception as e:
                pass
        # print(" done.")

    def show_main_window_event_handler(self):
        self.show()

    def hide_main_window_event_handler(self):
        self.btn_back_click()

    def sg_sync_signal_event_handler(self):
        print("Signal generator sync out detected")
        pass
    
    def interrupt_event_handler(self):
        if _platform == "linux" or _platform == "linux2":
            GPIO.remove_event_detect(signal_interrupt_input)
            print("Interrupt relay engage! Test in progress = " + str(self.is_test_in_progress))

    def interlock_event_handler(self):
        if _platform == "linux" or _platform == "linux2":
            sleep(0.1)
            self.is_lid_closed = (GPIO.input(signal_interlock_input) == 0)
            self.is_lid_open = (GPIO.input(signal_interlock_input) == 1)

            # self.is_lid_closed = self.interlock_signal.is_pressed
            # self.is_lid_open = not self.interlock_signal.is_pressed

            self.ui.btn_start.setVisible(self.is_lid_closed)
            if self.is_lid_closed:
                self.set_btn_color("btn_lbl_start", 1)
                self.ui.btn_lbl_start.setText("Press start")
            else:
                self.set_btn_color("btn_lbl_start", 5)
                self.ui.btn_lbl_start.setText("Lid open")
                self.backup_log_file()

            self.ui.statusbar.showMessage("")
            self.ui.statusbar.repaint()
            self.repaint()

    def btn_soft_start_click(self):
        self.is_soft_button = True
        self.btn_start_click()
        self.is_soft_button = False

    def btn_start_click(self):

        sleep(0.2)

        if _platform == "linux" or _platform == "linux2":
            self.is_lid_closed = (GPIO.input(signal_interlock_input) == 0)
            self.is_lid_open = (GPIO.input(signal_interlock_input) == 1)
            self.is_start_button_pressed = GPIO.input(btn_start_input) == 0

        self.is_start_button_released = not self.is_start_button_pressed

        # is Start button realy, realy pressed ?  Or GUI start button ?
        if self.is_start_button_released and (not self.is_soft_button):
            # should be 0 !!!
            return

        if self.test_in_progress:
            return

        if self.debug_mode:
            self.is_lid_open = False

        if self.is_lid_open:
            self.ui.statusbar.showMessage("LID is still open")
            self.ui.statusbar.repaint()
            sleep(1)
            self.ui.statusbar.showMessage("")
            self.ui.statusbar.repaint()
            return

        # No return beyond this point.
        self.test_in_progress = True

        if _platform == "linux" or _platform == "linux2":
            # Remove event detection in case of multiple press of start button
            GPIO.remove_event_detect(btn_start_input)

        self.repaint_gui()

        self.ui.statusbar.repaint()
        # Reset all test buttons to OFF grey color
        self.ui.btn_start.setEnabled(False)
        if self.debug_mode:
            # Don't read bar codes
            self.is_bar_code_on_slot = [True] * 4
        else:
            self.ui.btn_lbl_start.setText("Reading bar codes")
            self.set_btn_color("btn_lbl_start", 5)
            # Check presense of boards by reading bar codes or/and testing voltage
            self.is_bar_code_on_slot = [False] * 4
            self.check_dut_presense()

        # in case of debug edditing settings file, load it again
        # This line should be commented in production
        self.load_test_calibration_file()

        # Clear all previous test data results
        # self.test_pass_results = [True] * 30
        self.test_pass_result_slot_1 = [True] * 30
        self.test_pass_result_slot_2 = [True] * 30
        self.test_pass_result_slot_3 = [True] * 30
        self.test_pass_result_slot_4 = [True] * 30
        self.test_value_result_slot_1 = [0] * 30
        self.test_value_result_slot_2 = [0] * 30
        self.test_value_result_slot_3 = [0] * 30
        self.test_value_result_slot_4 = [0] * 30
        self.is_slot_1_test_pass = True
        self.is_slot_2_test_pass = True
        self.is_slot_3_test_pass = True
        self.is_slot_4_test_pass = True
        self.ui.btn_lbl_start.setText("Testing")
        self.set_btn_color("btn_lbl_start", 4)
        # And voila: Magic start
        for test in self.exec_test:
            self.test_number = test
            # file_name = self.product_name + ".t" + str(test)
            # Execute single test
            # Results :
            # 0 : completed successfuly
            # 1 : Communication error
            # 2 : Lid open during a test
            test_fail = 0
            test_fail = self.single_test()
            if test_fail:
                # User open a lid during a last test
                if _platform == "linux" or _platform == "linux2":
                    # Restore event detection for start button
                    GPIO.add_event_detect(btn_start_input, GPIO.RISING, callback=self.start_event_detected.emit, bouncetime=1000)
                self.test_in_progress = False
                self.repaint_gui()
                return
            else:
                dt = str(datetime.now())
                dt = dt[dt.find(" "):]
                dt = dt[:dt.find(".")]
                dt = dt[:len(dt) - 3]
                self.ui.lbl_date_time.setText(dt)
                # print("Screen repaint...", end="")
                self.repaint()
                # print("done.")

        # PASS / FAIL check
        all_unites_passed = True
        for slot_number in range(1, 5):
            if self.is_bar_code_on_slot[slot_number-1]:
                if not self.check_pass_fail(slot_number):
                    all_unites_passed = False
                self.units_tested += 1
        self.ui.lbl_units_tested.setText(str(self.units_tested))

        self.ui.btn_lbl_start.setText("TEST FINISHED")
        if all_unites_passed:
            self.set_btn_color("btn_lbl_start", 1)
        else:
            self.set_btn_color("btn_lbl_start", 2)

        # Full log file with all test results
        # self.save_log_file()

        self.save_new_format_log_file()

        self.loopCounter += 1
        self.ui.lcdLoop.display(self.loopCounter)
        self.ui.lcdLoop.repaint()
        self.ui.lbl_units_pass.setText(str(self.pass_units_counter_global))
        self.ui.lbl_units_fail.setText(str(self.fail_units_counter_global))
        self.ui.btn_start.repaint()
        if self.ui.cbLoop.isChecked():
            if self.loopCounter < self.ui.sbLoopCount.value():
                self.timer.start(1000)
                self.ui.btn_start.setEnabled(False)
            else:
                self.ui.btn_start.setEnabled(True)
                self.timer.stop()
        else:
            self.ui.btn_start.setEnabled(True)
            self.timer.stop()
        if _platform == "linux" or _platform == "linux2":
            # Restore event detection for start button
            GPIO.add_event_detect(btn_start_input, GPIO.RISING, callback=self.start_event_detected.emit, bouncetime=1000)
        self.test_in_progress = False

    def check_pass_fail(self, slot_number):
        # SLot 1 LED pass or fail ?
        # Hide labels "Last test" and "Measurements"
        exec("self.ui.lbl_lbl_test_no_" + str(slot_number) + ".setVisible(False)")
        exec("self.ui.lbl_test_no_" + str(slot_number) + ".setVisible(False)")
        exec("self.ui.lbl_lbl_measurement_" + str(slot_number) + ".setVisible(False)")
        exec("self.ui.lbl_measurement_" + str(slot_number) + ".setVisible(False)")
        unit_pass = True
        if slot_number == 1:
            for test_num in self.test_pass_result_slot_1:
                if not test_num:
                    unit_pass = False
                    self.is_slot_1_test_pass = False
                    break
        if slot_number == 2:
            for test_num in self.test_pass_result_slot_2:
                if not test_num:
                    unit_pass = False
                    self.is_slot_2_test_pass = False
                    break
        if slot_number == 3:
            for test_num in self.test_pass_result_slot_3:
                if not test_num:
                    unit_pass = False
                    self.is_slot_3_test_pass = False
                    break
        if slot_number == 4:
            for test_num in self.test_pass_result_slot_4:
                if not test_num:
                    unit_pass = False
                    all_pass = False
                    self.is_slot_4_test_pass = False
                    break
        exec("self.is_slot_" + str(slot_number) + "_test_pass = unit_pass")
        if unit_pass:
            self.pass_units_counter_global += 1
            pass_color = 1
            pass_fail_message = "PASS"
        else:
            self.fail_units_counter_global += 1
            pass_color = 2
            pass_fail_message = "FAIL"
        # Show colored rectangle on main screeen
        self.set_slot_pass_fail_color("fSlot_" + str(slot_number), pass_color)

        exec("self.ui.lbl_ResultTest_" + str(slot_number) + ".setText(pass_fail_message)")
        self.set_pf_leds(unit_pass, not unit_pass, slot_number)
        return unit_pass

    def set_pf_leds(self, led_p, led_f, slot_number):
        if self.debug_mode:
            return
        # Set pass and fail LED
        relay_led_p = self.what_is_ndb_relay_number("led_p")
        relay_led_f = self.what_is_ndb_relay_number("led_f")
        for err_count in range(1, 6):
            try:
                # req = eval("self.modbus_client.write_coil(relay_led_p, unit_pass, unit=ndb_" + str(slot_number) + ")")
                req = self.modbus_client.write_coil(relay_led_p, led_p, unit=eval("ndb_" + str(slot_number)))
                assert (not req.isError())
                break
            except Exception as e:
                log_error = "RON ndb_" + str(slot_number) + " LED PASS, attempt " + str(err_count)
                self.log_comm_error(log_error)
        for err_count in range(1, 6):
            try:
                # req = eval("self.modbus_client.write_coil(relay_led_f, not unit_pass, unit=ndb_" + str(slot_number) + ")")
                req = self.modbus_client.write_coil(relay_led_f, led_f, unit=eval("ndb_" + str(slot_number)))
                assert (not req.isError())
                break
            except Exception as e:
                log_error = "RON ndb_" + str(slot_number) + " LED FAIL, attempt " + str(err_count)
                self.log_comm_error(log_error)

    def cb_Loop_click(self):
        self.loopCounter = 0
        self.ui.lcdLoop.display(self.loopCounter)
        self.ui.lcdLoop.repaint()
        if self.ui.cbLoop.isChecked():
            self.ui.sbLoopCount.setEnabled(True)
            self.ui.sbLoopCount.repaint()
        else:
            self.ui.btn_start.setVisible(True)
            self.ui.sbLoopCount.setEnabled(False)
            self.ui.sbLoopCount.repaint()

    # ******************************************************************
    # ***************            TEST SEQUENCE           ***************
    # ******************************************************************
    def single_test(self):
        # Results :
        # 0 : completed successfuly
        # 1 : Communication error
        # 2 : Lid open during a test
        test_finish = 0
        comm_error = 1
        lid_open = 2
        test_result = test_finish
        print("Start test routine T" + str(self.test_number))
        # file = open(file_name, "r")
        # self.test_lines = [line.rstrip('\n') for line in file]
        # file.close()
        self.test_lines = self.tests[self.test_number-1].copy()

        # check for missing check start or check stop
        check_start_found = False
        check_loop_start_line_num = 0
        for line_num in range(0, len(self.test_lines)):
            line = self.test_lines[line_num].copy()
            command_param = str(line[1]).lower()
            if command_param == "start":
                check_start_found = False
                check_loop_start_line_num = line_num
                break

        check_stop_found = False
        check_loop_stop_line_num = 0
        for line_num in range(0, len(self.test_lines)):
            line = self.test_lines[line_num].copy()
            command_param = str(line[1]).lower()
            if command_param == "stop":
                check_stop_found = True
                check_loop_stop_line_num = line_num
                break

        if check_start_found and (not check_stop_found):
            QMessageBox.critical(self, "ERROR", "#  Check start command without stop line  #", QMessageBox.Ok)
            test_result = 3  # Command syntax error
            return test_result

        if check_stop_found and (not check_start_found):
            QMessageBox.critical(self, "ERROR", "#  Check stop command without start line  #", QMessageBox.Ok)
            test_result = 3  # Command syntax error
            return test_result

        line_num = 0
        self.slot = 1

        while line_num < len(self.test_lines):
            test_result = test_finish
            if _platform == "linux" or _platform == "linux2":
                if GPIO.input(signal_interlock_input) == 1:
                    test_result = lid_open
                    break
            line = self.test_lines[line_num].copy()
            line_num += 1
            # print("Exec line:" + str(line))
            print(str(line))
            command = str(line[0]).lower()
            if command == "check":
                command_param = str(line[1]).lower()
                if command_param == "start":
                    if not self.is_bar_code_on_slot[self.slot - 1]:
                        line = check_loop_stop_line_num
                if command_param == "stop":
                    if self.insert_number_duts <= self.slot:
                        self.slot += 1
                        line = check_loop_start_line_num
            if command == "slot":
                current_line_num = line_num
                # line_num = 0
                # Clear PASS / FAIL LEDS
                for pf_leds in range(1, 5):
                    if self.is_bar_code_on_slot[pf_leds - 1]:
                        self.set_pf_leds(False, False, pf_leds)
                self.slot = int(line[1])
                # Set both PASS/FAIL LED's on, so become orange
                # If there is bar code on slot, skip all lines until find another "Slot" or "End" line
                print("self.is_bar_code_on_slot  ", self.slot, self.is_bar_code_on_slot[self.slot - 1])
                if not self.is_bar_code_on_slot[self.slot - 1]:
                    # print("Change to new slot...")
                    while True:
                        line_num += 1
                        command_ = self.test_lines[line_num][0]
                        if command_ == "slot":
                            # line_num += 1
                            # print("New slot found on line:", line_num)
                            break
                        if command_ == "end":
                            # line_num += 1
                            # print("end found on line:", line_num)
                            break
                        if (line_num + 1) == len(self.test_lines):
                            # print("NO new slot line found")
                            line_num = current_line_num
                            break
                else:
                    self.set_pf_leds(True, True, self.slot)
            if command == "end":
                pass
            if command == "ron":
                test_result = self.exec_ron_command(line)
            if command == "roff":
                test_result = self.exec_roff_command(line)
            if command == "sg":
                self.exec_sg_command(line)
            if command == "mm":
                self.exec_mm_command(line)
            if command == "wait":
                self.exec_wait_command(line)
            if command == "pause":
                QMessageBox.information(self, "Test paused", "#       Press OK to continue       #", QMessageBox.Ok)
            if command == "pc":
                self.exec_pc_command(line)
            if command == "print":
                self.exec_print_command(line)
            if command == "msg":
                self.exec_msg_command(line)
            if command == "merr":
                if self.is_mm_error:
                    current_line_num = line_num
                    line_num = 0
                    label = ":" + parameters[1]
                    while True:
                        line_num += 1
                        line = self.test_lines[line_num].lstrip().replace(" ", "")
                        if line == label:
                            # print("Label '" + label + "' found on line " + str(line_num))
                            line_num += 1
                            break
                        if (line_num + 1) == len(self.test_lines):
                            # print("No '" + label + "' label found !!!")
                            line_num = current_line_num
                            break
            if command == "terr":
                if self.is_trip_error:
                    current_line_num = line_num
                    line_num = 0
                    label = ":" + parameters[1]
                    while True:
                        line_num += 1
                        line = self.test_lines[line_num].lstrip().replace(" ", "")
                        if line == label:
                            # print("Label '" + label + "' found on line " + str(line_num))
                            line_num += 1
                            break
                        if (line_num + 1) == len(self.test_lines):
                            # print("No '" + label + "' label found !!!")
                            line_num = current_line_num
                            break
            if command == "goto":
                current_line_num = line_num
                line_num = 0
                label = ":" + parameters[1]
                while True:
                    line_num += 1
                    line = self.test_lines[line_num].lstrip().replace(" ", "")
                    if line == label:
                        # print("Label '" + label + "' found on line " + str(line_num))
                        line_num += 1
                        break
                    if (line_num + 1) == len(self.test_lines):
                        # print("No '" + label + "' label found !!!")
                        line_num = current_line_num
                        break
            if command == "rsg":
                self.exec_rsg_commmand()
            if command == "trip":
                self.exec_trip_command(line)
            if command == "rbc":
                self.exec_rbc_command()
        return test_result

    def load_test_sequence_file(self):
        # Open products settings file
        wb = openpyxl.load_workbook(self.test_sequence_file, data_only=True)
        ps = wb["Settings"]
        self.test_sequence_file_issue = ps["B2"].value
        self.test_sequence_file_ref = ps["B3"].value
        self.test_specification_file = self.data_folder + ps["B5"].value
        self.test_calibration_file = self.data_folder + ps["B7"].value
        # Test_num
        # Voltage
        # Current
        # Execute
        self.test_sequence_lines = []
        self.exec_test = []
        for row_index in range(10, 50):
            test_sequence_line = []
            for cell in ps[row_index]:
                test_sequence_line.append(cell.value)
            if test_sequence_line[0] is not None:
                self.test_sequence_lines.append(test_sequence_line)
                if test_sequence_line[3] == '1':
                    self.exec_test.append(test_sequence_line[0])
        # Load tests
        # Commands
        # Parameters
        self.tests = []
        for test_num in self.test_sequence_lines:
            ps = wb["t"+str(test_num[0])]
            test_lines = []
            for row in ps:
                test_line = []
                for cell in row:
                    test_line.append(cell.value)
                if test_line[2] is not None and (str(test_line[0]).strip()[0] != "#"):
                    del test_line[0]    # remove disable column
                    del test_line[0]    # remove optional description column
                    test_lines.append(test_line)
            self.tests.append(test_lines)
        wb.close()
        # print(self.test_sequence_lines)
        # print(self.tests[0])

    def load_test_specification_file(self):
        # Open products specification file
        wb = openpyxl.load_workbook(self.test_specification_file, data_only=True)
        # Copy Specification sheet spec to test_specification list
        ps = wb["Specification"]
        self.test_specification_file_issue = ps["D1"].value
        self.test_specification_file_ref = ps["D2"].value
        self.test_specification_lines = []
        for row_index in range(8, 50):
            product_specification_line = []
            for cell in ps[row_index]:
                product_specification_line.append(cell.value)
            if product_specification_line[0] is not None:
                self.test_specification_lines.append(product_specification_line)
        wb.close()
        # print(self.test_specification_lines)

    def load_test_calibration_file(self):
        # Open products settings file
        wb = openpyxl.load_workbook(self.test_calibration_file, data_only=True)
        ps = wb["Calibration"]
        self.test_calibration_file_ref = ps["B2"].value
        # Test_num
        # SG Program
        # SG Setting
        # Frequency
        # Duration
        self.product_calibration_lines = []
        for row_index in range(6, 40):
            product_calibration_line = []
            for cell in ps[row_index]:
                product_calibration_line.append(cell.value)
            if product_calibration_line[0] is not None:
                self.product_calibration_lines.append(product_calibration_line)
        wb.close()
        # print(self.product_calibration_lines)

    """
        Check for board on slots by reading bar codes
    """
    def check_dut_presense(self):
        self.is_bar_code_on_slot = [False] * 4
        # print("Read all bar codes...")

        self.exec_rbc_command()
        print(self.barcode_slot_1)
        print(self.barcode_slot_2)
        print(self.barcode_slot_3)
        print(self.barcode_slot_4)

        #  BARCODE SLOT 1
        # print("BAR CODE SLOT 1 = ", self.barcode_slot_1)
        # If there is no bar code, read again
        if self.barcode_slot_1 == "000000000000000000":
            print("Read again bar code on slot 1")
            for bc_repeat in range(1, 3):
                self.barcode_slot_1 = self.read_bar_code(ndb_1)
                self.log_bar_code_error(1, bc_repeat)
                if self.barcode_slot_1 == "000000000000000000":
                    pass
                    # print("... no data")
                else:
                    # There is board on slot 1, and there is bar code. So board is valid
                    self.is_bar_code_on_slot[0] = True
                    print("self.barcode_slot_1:", self.barcode_slot_1)
                    break
        else:
            self.is_bar_code_on_slot[0] = True
        self.ui.lbl_bar_code_1.setText(self.barcode_slot_1)
        self.ui.lbl_bar_code_1.repaint()

        if self.insert_number_duts > 1:
            #  BARCODE SLOT 2
            # If there is no bar code, read again
            if self.barcode_slot_2 == "000000000000000000":
                print("Read again bar code on slot 2")
                for bc_repeat in range(1, 3):
                    self.barcode_slot_2 = self.read_bar_code(ndb_2)
                    # print("\n Read bar code again on slot 2, attemt " + str(bc_repeat), end="")
                    self.log_bar_code_error(2, bc_repeat)
                    if self.barcode_slot_2 == "000000000000000000":
                        pass
                        # # print(", no data")
                    else:
                        # There is board on slot 2 , and there is bar code. So board is valid
                        self.is_bar_code_on_slot[1] = True
                        print("self.barcode_slot_2:", self.barcode_slot_2)
                        break
            else:
                self.is_bar_code_on_slot[1] = True
            self.ui.lbl_bar_code_2.setText(self.barcode_slot_2)
            self.ui.lbl_bar_code_2.repaint()

        if self.insert_number_duts > 2:
            #  BARCODE SLOT 3
            # If there is no bar code, read again
            if self.barcode_slot_3 == "000000000000000000":
                print("Read again bar code on slot 3")
                for bc_repeat in range(1, 3):
                    self.barcode_slot_3 = self.read_bar_code(ndb_3)
                    # print("\n Read bar code again on slot 3, attemt " + str(bc_repeat), end="")
                    self.log_bar_code_error(3, bc_repeat)
                    if self.barcode_slot_3 == "000000000000000000":
                        pass
                        # print(", no data")
                    else:
                        # There is board on slot 3 , and there is bar code. So board is valid
                        self.is_bar_code_on_slot[2] = True
                        print("self.barcode_slot_3:", self.barcode_slot_3)
                        break
            else:
                self.is_bar_code_on_slot[2] = True
            self.ui.lbl_bar_code_3.setText(self.barcode_slot_3)
            self.ui.lbl_bar_code_3.repaint()

        if self.insert_number_duts > 3:
            #  BARCODE SLOT 4
            # If there is no bar code, read again
            if self.barcode_slot_4 == "000000000000000000":
                print("Read again bar code on slot 4")
                for bc_repeat in range(1, 3):
                    self.barcode_slot_4 = self.read_bar_code(ndb_4)
                    # print("\n Read bar code again on slot 4, attemt " + str(bc_repeat), end="")
                    self.log_bar_code_error(4, bc_repeat)
                    if self.barcode_slot_4 == "000000000000000000":
                        pass
                        # print(", no data")
                    else:
                        # There is board on slot 4 , and there is bar code. So board is valid
                        self.is_bar_code_on_slot[3] = True
                        print("self.barcode_slot_4:", self.barcode_slot_4)
                        break
            else:
                self.is_bar_code_on_slot[3] = True
            self.ui.lbl_bar_code_4.setText(self.barcode_slot_4)
            self.ui.lbl_bar_code_4.repaint()

        #  If still there is no bar code on some slot make clear notice
        if self.barcode_slot_1 == "000000000000000000":
            self.ui.lbl_lbl_test_no_1.setVisible(False)
            self.ui.lbl_test_no_1.setVisible(False)
            self.ui.lbl_lbl_measurement_1.setVisible(False)
            self.ui.lbl_measurement_1.setVisible(False)
            self.set_slot_pass_fail_color("fSlot_1", 3)
            self.ui.lbl_ResultTest_1.setText("ALL TESTS ABANDONED")
            self.ui.lbl_bar_code_1.setText("  BAR CODE ERROR  ")
        if self.insert_number_duts > 1:
            if self.barcode_slot_2 == "000000000000000000":
                self.ui.lbl_lbl_test_no_2.setVisible(False)
                self.ui.lbl_test_no_2.setVisible(False)
                self.ui.lbl_lbl_measurement_2.setVisible(False)
                self.ui.lbl_measurement_2.setVisible(False)
                self.set_slot_pass_fail_color("fSlot_2", 3)
                self.ui.lbl_ResultTest_2.setText("ALL TESTS ABANDONED")
                self.ui.lbl_bar_code_2.setText("  BAR CODE ERROR  ")
        else:
            self.set_slot_pass_fail_color("fSlot_2", 0)
            self.ui.lbl_ResultTest_2.setText("")
        if self.insert_number_duts > 2:
            if self.barcode_slot_3 == "000000000000000000":
                self.ui.lbl_lbl_test_no_3.setVisible(False)
                self.ui.lbl_test_no_3.setVisible(False)
                self.ui.lbl_lbl_measurement_3.setVisible(False)
                self.ui.lbl_measurement_3.setVisible(False)
                self.set_slot_pass_fail_color("fSlot_3", 3)
                self.ui.lbl_ResultTest_3.setText("ALL TESTS ABANDONED")
                self.ui.lbl_bar_code_3.setText("  BAR CODE ERROR  ")
        else:
            self.set_slot_pass_fail_color("fSlot_3", 0)
            self.ui.lbl_ResultTest_3.setText("")
        if self.insert_number_duts > 3:
            if self.barcode_slot_3 == "000000000000000000":
                # Hide labels "Last test" and "Measurements"
                self.ui.lbl_lbl_test_no_4.setVisible(False)
                self.ui.lbl_test_no_4.setVisible(False)
                self.ui.lbl_lbl_measurement_4.setVisible(False)
                self.ui.lbl_measurement_4.setVisible(False)
                self.set_slot_pass_fail_color("fSlot_4", 3)
                self.ui.lbl_ResultTest_4.setText("ALL TESTS ABANDONED")
                self.ui.lbl_bar_code_4.setText("  BAR CODE ERROR  ")
        else:
            self.set_slot_pass_fail_color("fSlot_4", 0)
            self.ui.lbl_ResultTest_4.setText("")

        self.is_enabled_log_mm_result = True

    def read_bar_code(self, board):
        for err_count in range(1, 3):
            try:
                req = self.modbus_client.write_coil(31, True, unit=board)
            except Exception as e:
                sleep(self.mbTimeout)
        sleep(1.5)
        for err_count in range(1, 3):
            try:
                req = self.modbus_client.read_holding_registers(2, 9, unit=board)
                assert (not req.isError())
                barcode = ""
                for register in range(0, 9):
                    hi = req.registers[register] >> 8
                    lo = req.registers[register] & 255
                    barcode = barcode + chr(lo) + chr(hi)
                break
            except Exception as e:
                barcode = "000000000000000000"
                sleep(self.mbTimeout)
        return str(barcode)

    def exec_rbc_command(self):
        for err_count in range(1, 3):
            try:
                # print(" Slot 1 send initial barcode read attempt " + str(err_count))
                req_slot_1 = self.modbus_client.write_coil(31, True, unit=ndb_1)
                assert (not req_slot_1.isError())

                if self.insert_number_duts > 1:
                    print(" Slot 2 send initial barcode read attempt " + str(err_count))
                    req_slot_2 = self.modbus_client.write_coil(31, True, unit=ndb_2)
                    assert (not req_slot_2.isError())

                if self.insert_number_duts > 2:
                    req_slot_3 = self.modbus_client.write_coil(31, True, unit=ndb_3)
                    assert (not req_slot_3.isError())

                if self.insert_number_duts > 3:
                    req_slot_4 = self.modbus_client.write_coil(31, True, unit=ndb_4)
                    assert (not req_slot_4.isError())
                break
            except Exception as e:
                print("Modbus Error sending inital read =", e)
                sleep(self.mbTimeout)
        sleep(1.5)
        for slot in range(1, self.insert_number_duts + 1):
            for err_count in range(1, 3):
                try:
                    # print(" Slot read barcode to slot " + str(err_count))
                    exec_str = "self.modbus_client.read_holding_registers(2, 9, unit=ndb_" + str(slot) + ")"
                    req = eval(exec_str)
                    assert (not req.isError())
                    barcode = ""
                    for register in range(0, 9):
                        hi = req.registers[register] >> 8
                        lo = req.registers[register] & 255
                        barcode = barcode + chr(lo) + chr(hi)
                    exec("self.barcode_slot_" + str(slot) + " = str(barcode)")
                    exec("self.ui.lbl_bar_code_" + str(slot) + ".setText(barcode)")
                    exec("self.ui.lbl_bar_code_" + str(slot) + ".repaint()")
                    break
                except Exception as e:
                    print("Modbus Error sending read barcode to slot =", str(slot))
                    print("Modbus Error =", e)
                    sleep(self.mbTimeout)

    def get_voltage_parameter(self):
        voltage = str(self.test_sequence_lines[self.test_number - 1][1]).lower()
        return voltage

    def get_current_parameter(self):
        current = str(self.test_sequence_lines[self.test_number - 1][2]).lower()
        return current

    def exec_ron_command(self, ron_command):
        if self.debug_mode:
            print("Debug mode, return from RON")
            return
        # Results :
        # 0 : completed successfuly
        # 1 : Communication error
        cmd_finish = 0
        comm_error = 1
        cmd_result = cmd_finish

        board = ron_command[1]
        board = ron_command[1]
        if board == "ndb":
            board = board + "_" + str(self.slot)

        relays = ron_command[2]
        parameters = relays.split(",")

        if board == "tmm":
            if len(parameters) < 4:
                for relay in parameters:
                    adr = int(relay) - 1
                    if adr < 3:
                        self.tmm_outputs[adr] = True
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tmm_outputs, unit=tmm)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("RON tmm, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "Multimeter board have only 3 relays, but command request is for " + str(len(parameters)) + " relays.")
        if board == "tmux_1":
            if len(parameters) < 9:
                for relay in parameters:
                    adr = int(relay) - 1
                    if adr < 9:
                        self.tmux_1_outputs[adr] = True
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tmux_1_outputs, unit=tmux_1)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("RON tmux_1, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "MUX boards have only 8 relays, but command request is for " + str(
                    len(parameters)) + " relays.")
        if board == "tmux_2":
            if len(parameters) < 9:
                if len(parameters) == 1:
                    relay_name = parameters[0].lower()
                    if relay_name == "dut":
                        relay_name = relay_name + str(self.slot)
                    if relay_name == "voff":
                        self.slot = 0
                        self.tmux_2_outputs[0] = False
                        self.tmux_2_outputs[1] = False
                        self.tmux_2_outputs[2] = False
                        self.tmux_2_outputs[3] = False
                        self.tmux_2_outputs[4] = False
                        self.tmux_2_outputs[5] = False
                        self.tmux_2_outputs[6] = False
                        self.tmux_2_outputs[7] = False
                    elif relay_name == "shunt":  # 1
                        self.slot = 0
                        self.tmux_2_outputs[0] = True
                        self.tmux_2_outputs[1] = False
                        self.tmux_2_outputs[2] = False
                        self.tmux_2_outputs[3] = False
                        self.tmux_2_outputs[4] = False
                        self.tmux_2_outputs[5] = False
                        self.tmux_2_outputs[6] = False
                        self.tmux_2_outputs[7] = False
                    elif relay_name == "dut1":  # 1, 5, 7
                        # self.slot = 1
                        self.tmux_2_outputs[0] = True
                        self.tmux_2_outputs[1] = False
                        self.tmux_2_outputs[2] = False
                        self.tmux_2_outputs[3] = False
                        self.tmux_2_outputs[4] = True
                        self.tmux_2_outputs[5] = False
                        self.tmux_2_outputs[6] = True
                        self.tmux_2_outputs[7] = False
                    elif relay_name == "dut2":  # 1, 3 ,4
                        # self.slot = 2
                        self.tmux_2_outputs[0] = True
                        self.tmux_2_outputs[1] = False
                        self.tmux_2_outputs[2] = True
                        self.tmux_2_outputs[3] = True
                        self.tmux_2_outputs[4] = False
                        self.tmux_2_outputs[5] = False
                        self.tmux_2_outputs[6] = False
                        self.tmux_2_outputs[7] = False
                    elif relay_name == "dut3":  # 1, 3, 4, 5, 7
                        # self.slot = 3
                        self.tmux_2_outputs[0] = True
                        self.tmux_2_outputs[1] = False
                        self.tmux_2_outputs[2] = True
                        self.tmux_2_outputs[3] = True
                        self.tmux_2_outputs[4] = True
                        self.tmux_2_outputs[5] = False
                        self.tmux_2_outputs[6] = True
                        self.tmux_2_outputs[7] = False
                    elif relay_name == "dut4":  # 1, 2
                        # self.slot = 4
                        self.tmux_2_outputs[0] = True
                        self.tmux_2_outputs[1] = True
                        self.tmux_2_outputs[2] = False
                        self.tmux_2_outputs[3] = False
                        self.tmux_2_outputs[4] = False
                        self.tmux_2_outputs[5] = False
                        self.tmux_2_outputs[6] = False
                        self.tmux_2_outputs[7] = False
                    elif relay_name in ["1", "2", "3", "4", "5", "6", "7", "8"]:
                        adr = int(relay_name) - 1
                        if adr < 16:
                            self.tmux_2_outputs[adr] = True
                    else:
                        for relay in parameters:
                            adr = int(relay) - 1
                            if adr < 16:
                                self.tmux_2_outputs[adr] = True
                else:
                    for relay in parameters:
                        adr = int(relay) - 1
                        if adr < 9:
                            self.tmux_2_outputs[adr] = True
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tmux_2_outputs, unit=tmux_2)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("RON tmux_2, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "MUX boards have only 8 relays, but command request is for " + str(
                    len(parameters)) + " relays.")
        if board == "tmux_3":
            if len(parameters) < 9:
                for relay in parameters:
                    adr = int(relay) - 1
                    if adr < 9:
                        self.tmux_3_outputs[adr] = True
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tmux_3_outputs, unit=tmux_3)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("RON tmux_3, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "MUX boards have only 8 relays, but command request is for " + str(
                    len(parameters)) + " relays.")
        if board == "tcb_1":
            if len(parameters) < 17:
                if len(parameters) == 1:
                    relay_name = parameters[0].lower()
                    if relay_name == "voltage" or relay_name == "v":
                        relay_name = self.get_voltage_parameter()
                    if relay_name == "current" or relay_name == "i":
                        relay_name = self.get_current_parameter()
                    print("TCB_1 Voltage = " + relay_name)
                    if relay_name == "v0":
                        self.tcb_1_outputs[0] = False
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = False
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = False
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = False
                        self.tcb_1_outputs[10] = False
                    elif relay_name == "v48":  # 1
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = False
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = False
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = False
                        self.tcb_1_outputs[10] = False
                    elif relay_name == "v48m":  # 1, 10, 11
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = False
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = False
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = True
                        self.tcb_1_outputs[10] = True
                    elif relay_name == "v55":  # 1, 5, 7, 8
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = True
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = True
                        self.tcb_1_outputs[7] = True
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = False
                        self.tcb_1_outputs[10] = False
                    elif relay_name == "v85":  # 1, 5, 7, 8, 9
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = True
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = True
                        self.tcb_1_outputs[7] = True
                        self.tcb_1_outputs[8] = True
                        self.tcb_1_outputs[9] = False
                        self.tcb_1_outputs[10] = False
                    elif relay_name == "v230":  # 1, 6
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = False
                        self.tcb_1_outputs[5] = True
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = False
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = False
                        self.tcb_1_outputs[10] = False
                    elif relay_name == "v230m":  # 1, 6, 11
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = False
                        self.tcb_1_outputs[5] = True
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = False
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = False
                        self.tcb_1_outputs[10] = True
                    elif relay_name == "v268l1":  # 1, 2, 5, 8
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = True
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = True
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = True
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = False
                        self.tcb_1_outputs[10] = False
                    elif relay_name == "v268l1m":  # 1, 2, 5, 8, 10, 11
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = True
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = True
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = True
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = True
                        self.tcb_1_outputs[10] = True
                    elif relay_name == "v268l2":  # 1, 3, 5, 8
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = True
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = True
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = True
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = False
                        self.tcb_1_outputs[10] = False
                    elif relay_name == "v268l2m":  # 1, 3, 5, 8, 10, 11
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = True
                        self.tcb_1_outputs[3] = False
                        self.tcb_1_outputs[4] = True
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = True
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = True
                        self.tcb_1_outputs[10] = True
                    elif relay_name == "v268l3":  # 1, 4, 5, 8
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = True
                        self.tcb_1_outputs[4] = True
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = True
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = False
                        self.tcb_1_outputs[10] = False
                    elif relay_name == "v268l3m":  # 1, 4, 5, 8, 10, 11
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = False
                        self.tcb_1_outputs[2] = False
                        self.tcb_1_outputs[3] = True
                        self.tcb_1_outputs[4] = True
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = True
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = True
                        self.tcb_1_outputs[10] = True
                    elif relay_name == "v268":  # 1, 2, 3, 4, 5, 8
                        self.tcb_1_outputs[0] = True
                        self.tcb_1_outputs[1] = True
                        self.tcb_1_outputs[2] = True
                        self.tcb_1_outputs[3] = True
                        self.tcb_1_outputs[4] = True
                        self.tcb_1_outputs[5] = False
                        self.tcb_1_outputs[6] = False
                        self.tcb_1_outputs[7] = True
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[8] = False
                        self.tcb_1_outputs[9] = False
                        self.tcb_1_outputs[10] = False
                    else:
                        for relay in parameters:
                            adr = int(relay) - 1
                            if adr < 16:
                                self.tcb_1_outputs[adr] = True
                else:
                    for relay in parameters:
                        adr = int(relay) - 1
                        if adr < 16:
                            self.tcb_1_outputs[adr] = True
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tcb_1_outputs, unit=tcb_1)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("RON tcb_1, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "Coil board have max 16 relays, but command request is for " + str(
                    len(parameters)) + " relays.")
        if board == "tcb_2":
            if len(parameters) < 17:
                if len(parameters) == 1:
                    relay_name = parameters[0].lower()
                    if relay_name == "voltage" or relay_name == "v":
                        relay_name = self.get_voltage_parameter()
                    if relay_name == "current" or relay_name == "i":
                        relay_name = self.get_current_parameter()
                    self.fc_setup = relay_name
                    print("TCB_2 Current = " + relay_name)
                    if relay_name == "fcoff" or relay_name == "I0":
                        self.tcb_2_outputs[0] = False
                        self.tcb_2_outputs[1] = False
                        self.tcb_2_outputs[2] = False
                        self.tcb_2_outputs[3] = False
                        self.tcb_2_outputs[4] = False
                        self.tcb_2_outputs[5] = False
                        self.tcb_2_outputs[6] = False
                        self.tcb_2_outputs[7] = False
                        self.tcb_2_outputs[8] = False
                        self.tcb_2_outputs[9] = False
                        self.tcb_2_outputs[10] = False
                        self.tcb_2_outputs[11] = False
                    elif relay_name == "i1":                    # 2,7
                        self.tcb_2_outputs[0] = False
                        self.tcb_2_outputs[1] = True
                        self.tcb_2_outputs[2] = False
                        self.tcb_2_outputs[3] = False
                        self.tcb_2_outputs[4] = False
                        self.tcb_2_outputs[5] = False
                        self.tcb_2_outputs[6] = True
                        self.tcb_2_outputs[7] = False
                        self.tcb_2_outputs[8] = False
                        self.tcb_2_outputs[9] = False
                        self.tcb_2_outputs[10] = False
                        self.tcb_2_outputs[11] = False
                    elif relay_name == "i2":                    # 2,4,7
                        self.tcb_2_outputs[0] = False
                        self.tcb_2_outputs[1] = True
                        self.tcb_2_outputs[2] = False
                        self.tcb_2_outputs[3] = True
                        self.tcb_2_outputs[4] = False
                        self.tcb_2_outputs[5] = False
                        self.tcb_2_outputs[6] = True
                        self.tcb_2_outputs[7] = False
                        self.tcb_2_outputs[8] = False
                        self.tcb_2_outputs[9] = False
                        self.tcb_2_outputs[10] = False
                        self.tcb_2_outputs[11] = False
                    elif relay_name == "i3":                    # 2,3,7
                        self.tcb_2_outputs[0] = False
                        self.tcb_2_outputs[1] = True
                        self.tcb_2_outputs[2] = True
                        self.tcb_2_outputs[3] = False
                        self.tcb_2_outputs[4] = False
                        self.tcb_2_outputs[5] = False
                        self.tcb_2_outputs[6] = True
                        self.tcb_2_outputs[7] = False
                        self.tcb_2_outputs[8] = False
                        self.tcb_2_outputs[9] = False
                        self.tcb_2_outputs[10] = False
                        self.tcb_2_outputs[11] = False
                    elif relay_name == "i4":                    # 1, 5, 7, 8
                        self.tcb_2_outputs[0] = True
                        self.tcb_2_outputs[1] = False
                        self.tcb_2_outputs[2] = False
                        self.tcb_2_outputs[3] = False
                        self.tcb_2_outputs[4] = True
                        self.tcb_2_outputs[5] = False
                        self.tcb_2_outputs[6] = True
                        self.tcb_2_outputs[7] = True
                        self.tcb_2_outputs[8] = False
                        self.tcb_2_outputs[9] = False
                        self.tcb_2_outputs[10] = False
                        self.tcb_2_outputs[11] = False
                    elif relay_name == "i5":                    # 1, 5, 7, 8, 11
                        self.tcb_2_outputs[0] = True
                        self.tcb_2_outputs[1] = False
                        self.tcb_2_outputs[2] = False
                        self.tcb_2_outputs[3] = False
                        self.tcb_2_outputs[4] = True
                        self.tcb_2_outputs[5] = False
                        self.tcb_2_outputs[6] = True
                        self.tcb_2_outputs[7] = True
                        self.tcb_2_outputs[8] = False
                        self.tcb_2_outputs[9] = False
                        self.tcb_2_outputs[10] = True
                        self.tcb_2_outputs[11] = False
                    elif relay_name == "i6":                    # 1, 5, 7, 8, 10
                        self.tcb_2_outputs[0] = True
                        self.tcb_2_outputs[1] = False
                        self.tcb_2_outputs[2] = False
                        self.tcb_2_outputs[3] = False
                        self.tcb_2_outputs[4] = True
                        self.tcb_2_outputs[5] = False
                        self.tcb_2_outputs[6] = True
                        self.tcb_2_outputs[7] = True
                        self.tcb_2_outputs[8] = False
                        self.tcb_2_outputs[9] = True
                        self.tcb_2_outputs[10] = False
                        self.tcb_2_outputs[11] = False
                    elif relay_name == "i7":                    # 1, 5, 6, 7, 8, 9
                        self.tcb_2_outputs[0] = True
                        self.tcb_2_outputs[1] = False
                        self.tcb_2_outputs[2] = False
                        self.tcb_2_outputs[3] = False
                        self.tcb_2_outputs[4] = True
                        self.tcb_2_outputs[5] = True
                        self.tcb_2_outputs[6] = True
                        self.tcb_2_outputs[7] = True
                        self.tcb_2_outputs[8] = True
                        self.tcb_2_outputs[9] = False
                        self.tcb_2_outputs[10] = False
                        self.tcb_2_outputs[11] = False
                    elif relay_name == "i8":                    # 1, 5, 6, 7, 8, 12
                        self.tcb_2_outputs[0] = True
                        self.tcb_2_outputs[1] = False
                        self.tcb_2_outputs[2] = False
                        self.tcb_2_outputs[3] = False
                        self.tcb_2_outputs[4] = True
                        self.tcb_2_outputs[5] = True
                        self.tcb_2_outputs[6] = True
                        self.tcb_2_outputs[7] = True
                        self.tcb_2_outputs[8] = False
                        self.tcb_2_outputs[9] = False
                        self.tcb_2_outputs[10] = False
                        self.tcb_2_outputs[11] = True
                    else:
                        for relay in parameters:
                            adr = int(relay) - 1
                            if adr < 16:
                                self.tcb_1_outputs[adr] = True
                else:
                    for relay in parameters:
                        adr = int(relay) - 1
                        if adr < 16:
                            self.tcb_1_outputs[adr] = True
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tcb_2_outputs, unit=tcb_2)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("RON tcb_2, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "Coil board have max 16 relays, but command request is for " + str(
                    len(parameters)) + " relays.")
        if board == "tcb_3":
            if len(parameters) < 17:
                if len(parameters) == 1:
                    self.fc_setup = ""
                    relay_name = parameters[0].lower()
                    if relay_name == "voltage" or relay_name == "v":
                        relay_name = self.get_voltage_parameter()
                    if relay_name == "current" or relay_name == "i":
                        relay_name = self.get_current_parameter()
                    print("TCB_3 Voltage/Current = " + relay_name)
                    if relay_name == "v0":
                        self.tcb_3_outputs[0] = False
                        self.tcb_3_outputs[1] = False
                        self.tcb_3_outputs[2] = False
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = False
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = False
                        self.tcb_3_outputs[10] = False
                        self.tcb_3_outputs[11] = False
                        self.tcb_3_outputs[12] = False
                        self.tcb_3_outputs[13] = False
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = False
                    elif relay_name == "10.7":
                        # 3, 5
                        self.tcb_3_outputs[0] = False
                        self.tcb_3_outputs[1] = False
                        self.tcb_3_outputs[2] = True
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = True
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = False
                        self.tcb_3_outputs[10] = False
                        self.tcb_3_outputs[11] = False
                        self.tcb_3_outputs[12] = False
                        self.tcb_3_outputs[13] = False
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = False
                    elif relay_name == "13.5":
                        # 1, 3, 5
                        self.tcb_3_outputs[0] = True
                        self.tcb_3_outputs[1] = False
                        self.tcb_3_outputs[2] = True
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = True
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = False
                        self.tcb_3_outputs[10] = False
                        self.tcb_3_outputs[11] = False
                        self.tcb_3_outputs[12] = False
                        self.tcb_3_outputs[13] = False
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = False
                    elif relay_name == "26.4":
                        # 1, 2, 3, 5
                        self.tcb_3_outputs[0] = True
                        self.tcb_3_outputs[1] = True
                        self.tcb_3_outputs[2] = True
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = True
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = False
                        self.tcb_3_outputs[10] = False
                        self.tcb_3_outputs[11] = False
                        self.tcb_3_outputs[12] = False
                        self.tcb_3_outputs[13] = False
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = False
                    elif relay_name == "10.7m":
                        # 5, 14, 16
                        self.tcb_3_outputs[0] = False
                        self.tcb_3_outputs[1] = False
                        self.tcb_3_outputs[2] = False
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = True
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = False
                        self.tcb_3_outputs[10] = False
                        self.tcb_3_outputs[11] = False
                        self.tcb_3_outputs[12] = False
                        self.tcb_3_outputs[13] = True
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = True
                    elif relay_name == "13.5m":
                        # 1, 5, 14, 16
                        self.tcb_3_outputs[0] = True
                        self.tcb_3_outputs[1] = False
                        self.tcb_3_outputs[2] = False
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = True
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = False
                        self.tcb_3_outputs[10] = False
                        self.tcb_3_outputs[11] = False
                        self.tcb_3_outputs[12] = False
                        self.tcb_3_outputs[13] = True
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = True
                    elif relay_name == "26.4m":
                        # 1, 2, 5, 14, 16
                        self.tcb_3_outputs[0] = True
                        self.tcb_3_outputs[1] = True
                        self.tcb_3_outputs[2] = False
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = True
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = False
                        self.tcb_3_outputs[10] = False
                        self.tcb_3_outputs[11] = False
                        self.tcb_3_outputs[12] = False
                        self.tcb_3_outputs[13] = True
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = True
                    elif relay_name == "i1":
                        # 10, 16
                        self.fc_setup = relay_name
                        self.tcb_3_outputs[0] = False
                        self.tcb_3_outputs[1] = False
                        self.tcb_3_outputs[2] = False
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = False
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = True
                        self.tcb_3_outputs[10] = False
                        self.tcb_3_outputs[11] = False
                        self.tcb_3_outputs[12] = False
                        self.tcb_3_outputs[13] = False
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = True
                    elif relay_name == "i2":
                        # 10, 13, 16
                        self.fc_setup = relay_name
                        self.tcb_3_outputs[0] = False
                        self.tcb_3_outputs[1] = False
                        self.tcb_3_outputs[2] = False
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = False
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = True
                        self.tcb_3_outputs[10] = False
                        self.tcb_3_outputs[11] = False
                        self.tcb_3_outputs[12] = True
                        self.tcb_3_outputs[13] = False
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = True
                    elif relay_name == "i3":
                        # 10, 12, 16
                        self.fc_setup = relay_name
                        self.tcb_3_outputs[0] = False
                        self.tcb_3_outputs[1] = False
                        self.tcb_3_outputs[2] = False
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = False
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = True
                        self.tcb_3_outputs[10] = False
                        self.tcb_3_outputs[11] = True
                        self.tcb_3_outputs[12] = False
                        self.tcb_3_outputs[13] = False
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = True
                    elif relay_name == "i4":
                        # 10, 11, 16
                        self.fc_setup = relay_name
                        self.tcb_3_outputs[0] = False
                        self.tcb_3_outputs[1] = False
                        self.tcb_3_outputs[2] = False
                        self.tcb_3_outputs[3] = False
                        self.tcb_3_outputs[4] = False
                        self.tcb_3_outputs[5] = False
                        self.tcb_3_outputs[6] = False
                        self.tcb_3_outputs[7] = False
                        self.tcb_3_outputs[8] = False
                        self.tcb_3_outputs[9] = True
                        self.tcb_3_outputs[10] = True
                        self.tcb_3_outputs[11] = False
                        self.tcb_3_outputs[12] = False
                        self.tcb_3_outputs[13] = False
                        self.tcb_3_outputs[14] = False
                        self.tcb_3_outputs[15] = True
                    else:
                        for relay in parameters:
                            adr = int(relay) - 1
                            if adr < 16:
                                self.tcb_3_outputs[adr] = True
                else:
                    for relay in parameters:
                        adr = int(relay) - 1
                        if adr < 16:
                            self.tcb_3_outputs[adr] = True
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tcb_3_outputs, unit=tcb_3)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("RON tcb_3, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "Coil board 3 have max 16 relays, but command request is for " + str(
                    len(parameters)) + " relays.")
        if board == "ndb_1":
            # self.slot = 1
            if len(parameters) == 1:
                relay = self.what_is_ndb_relay_number(parameters[0])
                for err_count in range(1, 10):
                    try:
                        req = self.modbus_client.write_coil(relay, True, unit=ndb_1)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        log_error = "RON ndb_1, test:" + str(self.test_number) + ", attempt " + str(err_count)
                        self.log_comm_error(log_error)
            else:
                self.ndb_1_outputs = [False] * 24
                for relay in parameters:
                    adr = self.what_is_ndb_relay_number(relay) - 1
                    self.ndb_1_outputs[adr] = True
                # print(self.ndb_1_outputs)
                for err_count in range(1, 10):
                    try:
                        req = self.modbus_client.write_coils(0, self.ndb_1_outputs, unit=ndb_1)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        log_error = "RON ndb_1, test:" + str(self.test_number) + ", attempt " + str(err_count)
                        self.log_comm_error(log_error)
        if board == "ndb_2":
            # self.slot = 2
            if len(parameters) == 1:
                relay = self.what_is_ndb_relay_number(parameters[0])
                for err_count in range(1, 10):
                    try:
                        req = self.modbus_client.write_coil(relay, True, unit=ndb_2)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        log_error = "RON ndb_2, test:" + str(self.test_number) + ", attempt " + str(err_count)
                        self.log_comm_error(log_error)
            else:
                self.ndb_2_outputs = [False] * 24
                for relay in parameters:
                    adr = self.what_is_ndb_relay_number(relay)
                    self.ndb_2_outputs[adr] = True
                for err_count in range(1, 10):
                    try:
                        req = self.modbus_client.write_coils(0, self.ndb_2_outputs, unit=ndb_2)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        log_error = "RON ndb_2, test:" + str(self.test_number) + ", attempt " + str(err_count)
                        self.log_comm_error(log_error)
        if board == "ndb_3":
            # self.slot = 3
            if len(parameters) == 1:
                relay = self.what_is_ndb_relay_number(parameters[0])
                for err_count in range(1, 10):
                    try:
                        req = self.modbus_client.write_coil(relay, True, unit=ndb_3)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        log_error = "RON ndb_3, test:" + str(self.test_number) + ", attempt " + str(err_count)
                        self.log_comm_error(log_error)
            else:
                self.ndb_3_outputs = [False] * 24
                for relay in parameters:
                    adr = self.what_is_ndb_relay_number(relay)
                    self.ndb_3_outputs[adr] = True
                for err_count in range(1, 10):
                    try:
                        req = self.modbus_client.write_coils(0, self.ndb_3_outputs, unit=ndb_3)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        log_error = "RON ndb_3, test:" + str(self.test_number) + ", attempt " + str(err_count)
                        self.log_comm_error(log_error)
        if board == "ndb_4":
            # self.slot = 4
            if len(parameters) == 1:
                relay = self.what_is_ndb_relay_number(parameters[0])
                for err_count in range(1, 10):
                    try:
                        req = self.modbus_client.write_coil(relay, True, unit=ndb_4)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        log_error = "RON ndb_4, test:" + str(self.test_number) + ", attempt " + str(err_count)
                        self.log_comm_error(log_error)
            else:
                self.ndb_4_outputs = [False] * 24
                for relay in parameters:
                    adr = self.what_is_ndb_relay_number(relay)
                    self.ndb_4_outputs[adr] = True
                for err_count in range(1, 10):
                    try:
                        req = self.modbus_client.write_coils(0, self.ndb_4_outputs, unit=ndb_4)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        log_error = "RON ndb_4, test:" + str(self.test_number) + ", attempt " + str(err_count)
                        self.log_comm_error(log_error)
        return cmd_result

    def exec_roff_command(self, roff_command):
        if self.debug_mode:
            print("Debug mode, return from ROFF")
            return
        # Results :
        # 0 : completed successfuly
        # 1 : Communication error
        cmd_finish = 0
        comm_error = 1
        cmd_result = cmd_finish

        board = roff_command[1]
        board = roff_command[1]
        if board == "ndb":
            board = board + "_" + str(self.slot)

        relays = roff_command[2]
        parameters = relays.split(",")

        if board == "tmm":
            if len(parameters) < 4:
                if parameters[0] == "all":
                    self.tmm_outputs = [False] * 3
                else:
                    for relay in parameters:
                        self.tmm_outputs[int(relay) - 1] = False
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tmm_outputs, unit=tmm)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("ROFF tmm, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "Multimeter board have only 3 relays, but command request is for " + str(
                    len(parameters)) + " relays.")
        if board == "tmux_1":
            if len(parameters) < 9:
                if parameters[0] == "all":
                    self.tmux_1_outputs = [False] * 8
                else:
                    for relay in parameters:
                        self.tmux_1_outputs[int(relay) - 1] = False
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tmux_1_outputs, unit=tmux_1)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("ROFF tmux_1, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "MUX boards have only 8 relays, but command request is for " + str(
                    len(parameters)) + " relays.")
        if board == "tmux_2":
            if len(parameters) < 9:
                if parameters[0] == "all":
                    self.tmux_2_outputs = [False] * 8
                else:
                    for relay in parameters:
                        self.tmux_2_outputs[int(relay) - 1] = False
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tmux_2_outputs, unit=tmux_2)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("ROFF tmux_2, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "MUX boards have only 8 relays, but command request is for " + str(
                    len(parameters)) + " relays.")
        if board == "tmux_3":
            if len(parameters) < 9:
                if parameters[0] == "all":
                    self.tmux_3_outputs = [False] * 8
                else:
                    for relay in parameters:
                        self.tmux_3_outputs[int(relay) - 1] = False
                for err_count in range(1, 6):
                    try:
                        req = self.modbus_client.write_coils(0, self.tmux_3_outputs, unit=tmux_3)
                        assert (not req.isError())
                        cmd_result = cmd_finish
                        break
                    except Exception as e:
                        sleep(self.mbTimeout)
                        cmd_result = comm_error
                        self.log_comm_error("ROFF tmux_3, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "MUX boards have only 8 relays, but command request is for " + str(
                    len(parameters)) + " relays.")
        if board == "tcb_1":
            if len(parameters) < 17:
                if len(parameters) == 1:
                    relay_name = parameters[0].lower()
                    if relay_name == "voltage" or relay_name == "v":
                        relay_name = self.get_voltage_parameter()
                    if relay_name == "current" or relay_name == "i":
                        relay_name = self.get_current_parameter()
                    if relay_name == "all":
                        self.tcb_1_outputs = [False] * 16
                    else:
                        for relay in parameters:
                            adr = int(relay) - 1
                            if adr < 16:
                                self.tcb_1_outputs[adr] = False
                    for err_count in range(1, 6):
                        try:
                            req = self.modbus_client.write_coils(0, self.tcb_1_outputs, unit=tcb_1)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF tcb_1, test:" + str(self.test_number) + ", attempt " + str(err_count))
                else:
                    for relay in parameters:
                        adr = int(relay) - 1
                        if adr < 16:
                            self.tcb_1_outputs[adr] = True
                    for err_count in range(1, 6):
                        try:
                            req = self.modbus_client.write_coils(0, self.tcb_1_outputs, unit=tcb_1)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF tcb_1, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "Coil board have max 16 relays, request is for " + str(len(parameters)) + " relays.")
        if board == "tcb_2":
            if len(parameters) < 17:
                if len(parameters) == 1:
                    relay_name = parameters[0].lower()
                    if relay_name == "voltage" or relay_name == "v":
                        relay_name = self.get_voltage_parameter()
                    if relay_name == "current" or relay_name == "i":
                        relay_name = self.get_current_parameter()
                    if relay_name == "all":
                        self.tcb_2_outputs = [False] * 16
                    else:
                        for relay in parameters:
                            adr = int(relay) - 1
                            if adr < 16:
                                self.tcb_2_outputs[adr] = False
                    for err_count in range(1, 6):
                        try:
                            req = self.modbus_client.write_coils(0, self.tcb_2_outputs, unit=tcb_2)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF tcb_2, test:" + str(self.test_number) + ", attempt " + str(err_count))
                else:
                    for relay in parameters:
                        adr = int(relay) - 1
                        if adr < 16:
                            self.tcb_2_outputs[adr] = True
                        for err_count in range(1, 6):
                            try:
                                req = self.modbus_client.write_coils(0, self.tcb_2_outputs, unit=tcb_2)
                                assert (not req.isError())
                                cmd_result = cmd_finish
                                break
                            except Exception as e:
                                sleep(self.mbTimeout)
                                cmd_result = comm_error
                                self.log_comm_error("ROFF tcb_2, test:" +
                                                    str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "Coil board have max 16 relays, request is for " + str(len(parameters)) + " relays.")
        if board == "tcb_3":
            if len(parameters) < 17:
                if len(parameters) == 1:
                    relay_name = parameters[0].lower()
                    if relay_name == "voltage" or relay_name == "v":
                        relay_name = self.get_voltage_parameter()
                    if relay_name == "current" or relay_name == "i":
                        relay_name = self.get_current_parameter()
                    if relay_name == "all":
                        self.tcb_3_outputs = [False] * 16
                    else:
                        for relay in parameters:
                            adr = int(relay) - 1
                            if adr < 16:
                                self.tcb_3_outputs[adr] = False
                    for err_count in range(1, 6):
                        try:
                            req = self.modbus_client.write_coils(0, self.tcb_3_outputs, unit=tcb_3)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF tcb_3, test:" + str(self.test_number) + ", attempt " + str(err_count))
                else:
                    for relay in parameters:
                        adr = int(relay) - 1
                        if adr < 16:
                            self.tcb_3_outputs[adr] = True
                        for err_count in range(1, 6):
                            try:
                                req = self.modbus_client.write_coils(0, self.tcb_3_outputs, unit=tcb_3)
                                assert (not req.isError())
                                cmd_result = cmd_finish
                                break
                            except Exception as e:
                                sleep(self.mbTimeout)
                                cmd_result = comm_error
                                self.log_comm_error("ROFF tcb_3, test:" +
                                                    str(self.test_number) + ", attempt " + str(err_count))
            else:
                QMessageBox.information(self, "Value error", "Coil board 3 have max 16 relays, request is for " + str(len(parameters)) + " relays.")
        if board == "ndb_1":
            # self.slot = 1
            if len(parameters) == 1:
                relay_name = parameters[0]
                if relay_name == "all":
                    self.ndb_1_outputs = [False] * 24
                    for err_count in range(1, 2):
                        try:
                            req = self.modbus_client.write_coils(0, self.ndb_1_outputs, unit=ndb_1)
                            assert (not req.isError())
                            break
                        except Exception as e:
                            self.log_comm_error("ROFF ndb_1, test:" + str(self.test_number) + ", attempt " + str(err_count))
                else:
                    relay = self.what_is_ndb_relay_number(relay_name)
                    for err_count in range(1, 10):
                        try:
                            req = self.modbus_client.write_coil(relay, False, unit=ndb_1)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF ndb_1, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                for relay in parameters:
                    adr = self.what_is_ndb_relay_number(relay) - 1
                    self.ndb_1_outputs[adr] = False
                    for err_count in range(1, 6):
                        try:
                            req = self.modbus_client.write_coils(0, self.ndb_1_outputs, unit=ndb_1)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF ndb_1, test:" + str(self.test_number) + ", attempt " + str(err_count))
        if board == "ndb_2":
            # self.slot = 2
            if len(parameters) == 1:
                relay_name = parameters[0]
                if relay_name == "all":
                    self.ndb_2_outputs = [False] * 24
                    for err_count in range(1, 10):
                        try:
                            req = self.modbus_client.write_coils(0, self.ndb_2_outputs, unit=ndb_2)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF ndb_2, test:" + str(self.test_number) + ", attempt " + str(err_count))
                else:
                    relay = self.what_is_ndb_relay_number(relay_name)
                    for err_count in range(1, 6):
                        try:
                            req = self.modbus_client.write_coil(relay, False, unit=ndb_2)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF ndb_2, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                for relay in parameters:
                    adr = self.what_is_ndb_relay_number(relay) - 1
                    self.ndb_2_outputs[adr] = False
                    for err_count in range(1, 10):
                        try:
                            req = self.modbus_client.write_coils(0, self.ndb_2_outputs, unit=ndb_2)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF ndb_2, test:" + str(self.test_number) + ", attempt " + str(err_count))
        if board == "ndb_3":
            # self.slot = 3
            if len(parameters) == 1:
                relay_name = parameters[0]
                if relay_name == "all":
                    self.ndb_3_outputs = [False] * 24
                    for err_count in range(1, 10):
                        try:
                            req = self.modbus_client.write_coils(0, self.ndb_3_outputs, unit=ndb_3)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF ndb_3, test:" + str(self.test_number) + ", attempt " + str(err_count))
                else:
                    relay = self.what_is_ndb_relay_number(relay_name)
                    for err_count in range(1, 10):
                        try:
                            req = self.modbus_client.write_coil(relay, False, unit=ndb_3)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF ndb_3, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                for relay in parameters:
                    adr = self.what_is_ndb_relay_number(relay) - 1
                    self.ndb_3_outputs[adr] = False
                    for err_count in range(1, 10):
                        try:
                            req = self.modbus_client.write_coils(0, self.ndb_3_outputs, unit=ndb_3)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF ndb_3, test:" + str(self.test_number) + ", attempt " + str(err_count))
        if board == "ndb_4":
            # self.slot = 4
            if len(parameters) == 1:
                relay_name = parameters[0]
                if relay_name == "all":
                    self.ndb_4_outputs = [False] * 24
                    for err_count in range(1, 10):
                        try:
                            req = self.modbus_client.write_coils(0, self.ndb_4_outputs, unit=ndb_4)
                            assert (not req.isError())
                            break
                        except Exception as e:
                            self.log_comm_error("ROFF ndb_4, test:" + str(self.test_number) + ", attempt " + str(err_count))
                else:
                    relay = self.what_is_ndb_relay_number(relay_name)
                    for err_count in range(1, 10):
                        try:
                            req = self.modbus_client.write_coil(relay, False, unit=ndb_4)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF ndb_4, test:" + str(self.test_number) + ", attempt " + str(err_count))
            else:
                for relay in parameters:
                    adr = self.what_is_ndb_relay_number(relay) - 1
                    self.ndb_4_outputs[adr] = False
                    for err_count in range(1, 10):
                        try:
                            req = self.modbus_client.write_coils(0, self.ndb_4_outputs, unit=ndb_4)
                            assert (not req.isError())
                            cmd_result = cmd_finish
                            break
                        except Exception as e:
                            sleep(self.mbTimeout)
                            cmd_result = comm_error
                            self.log_comm_error("ROFF ndb_4, test:" + str(self.test_number) + ", attempt " + str(err_count))
        return cmd_result

    def exec_sg_command(self, sg_command):
        # remove leading "sg" command string
        del sg_command[0]

        sg_program = sg_command[0]
        if sg_program == "run":
            sg_program = self.sg_calibration(1)
            sg_command[1] = self.sg_calibration(3)
            sg_command[2] = self.sg_calibration(4)
        freq = 0
        duration = 0
        p1 = 0
        p2 = 0
        p3 = 0
        p4 = 0
        p5 = 0
        print("Execute SG program " + sg_program)
        if sg_program == "21":
            # AC signal, Trip Time
            # Idn AC
            freq = float(sg_command[1])  # .encode('UTF-8')
            duration = float(sg_command[2]) / 1000  # in seconds
            amplitude = self.sg_settings(sg_program, 1)
            one_period = float(1 / freq)
            cycles = round(duration / one_period)
            self.sg_cmd = 'WAVE SINE\nFREQ ' + str(freq) + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nDCOFFS 0\nBST NCYC\nBSTCOUNT ' + str(cycles) + '\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        if sg_program == "22":
            # AC signal, Trip Time
            # Pulsating DC 7Idn
            freq = float(sg_command[1])  # .encode('UTF-8')
            duration = float(sg_command[2]) / 1000  # in seconds
            amplitude = self.sg_settings(sg_program, 1)
            one_period = float(1 / freq)
            cycles = round(duration / one_period)  # .encode('UTF-8')
            self.sg_cmd = 'ARBLOAD ARB1\nDCOFFS 0\nFREQ ' + str(freq) + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nBST NCYC\nBSTCOUNT ' + str(cycles) + '\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        if sg_program == "23":
            #  DC signal (10A), Trip Time
            amplitude = self.sg_settings(sg_program, 1)
            dcofss = str(float(amplitude) - 0.01)
            self.sg_cmd = 'WAVE PULSE\nDCOFFS ' + dcofss + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nPULSFREQ 9\nPULSSYMM 60\nPULSEDGE 0.00000001\nBST NCYC\nBSTCOUNT 1\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        if sg_program == "24a":
            # AC signal, Trip/No trip
            # LL AC
            freq = float(sg_command[1])
            duration = float(sg_command[2]) / 1000  # in seconds
            amplitude = self.sg_settings(sg_program, 1)
            one_period = float(1 / freq)
            cycles = round(duration / one_period)  # .encode('UTF-8')
            self.sg_cmd = 'WAVE SINE\nFREQ ' + str(freq) + '\nDCOFFS 0\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nBST NCYC\nBSTCOUNT ' + str(cycles) + '\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        if sg_program == "24b":
            # AC signal, Trip/No trip
            # UL AC
            freq = float(sg_command[1])
            duration = float(sg_command[2]) / 1000  # in seconds
            amplitude = self.sg_settings(sg_program, 1)
            one_period = float(1 / freq)
            cycles = round(duration / one_period)
            self.sg_cmd = 'WAVE SINE\nFREQ ' + str(freq) + '\nDCOFFS 0\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nBST NCYC\nBSTCOUNT ' + str(cycles) + '\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        if sg_program == "25a":
            # DC signal, Trip/no trip
            # LL DC I
            amplitude = self.sg_settings(sg_program, 1)
            dcofss = str(float(amplitude) - 0.01)
            self.sg_cmd = 'WAVE PULSE\nDCOFFS ' + dcofss + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nPULSFREQ 0.9\nPULSSYMM 95\nPULSEDGE 0.00000001\nBST NCYC\nBSTCOUNT 1\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        if sg_program == "25b":
            # DC signal, Trip/no trip
            # UL DC I
            amplitude = self.sg_settings(sg_program, 1)
            dc_offset = str(float(amplitude) - 0.01)
            self.sg_cmd = 'WAVE PULSE\nDCOFFS ' + dc_offset + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nPULSFREQ 0.9\nPULSSYMM 95\nPULSEDGE 0.00000001\nBST NCYC\nBSTCOUNT 1\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        if sg_program == "25c":
            # DC signal, Trip/no trip
            # LL DC II
            amplitude = self.sg_settings(sg_program, 1)
            dc_offset = str(float(amplitude) - 0.01)
            self.sg_cmd = 'WAVE PULSE\nDCOFFS -' + dc_offset + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nPULSFREQ 0.9\nPULSSYMM 95\nPULSEDGE 0.00000001\nBST NCYC\nBSTCOUNT 1\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT INVERT\nOUTPUT ON\n*TRG\n'
        if sg_program == "25d":
            # DC signal, Trip/no trip
            # UL DC II
            amplitude = self.sg_settings(sg_program, 1)
            dc_offset = str(float(amplitude) - 0.01)
            self.sg_cmd = 'WAVE PULSE\nDCOFFS -' + dc_offset + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nPULSFREQ 0.9\nPULSSYMM 95\nPULSEDGE 0.00000001\nBST NCYC\nBSTCOUNT 1\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT INVERT\nOUTPUT ON\n*TRG\n'
        if sg_program == "26":
            # AC 1kHz, Trip Time
            freq = float(sg_command[1])
            duration = float(sg_command[2]) / 1000  # in seconds
            amplitude = self.sg_settings(sg_program, 1)
            one_period = float(1 / freq)
            cycles = round(duration / one_period)
            self.sg_cmd = 'WAVE SINE\nFREQ ' + str(freq) + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nDCOFFS 0\nBST NCYC\nBSTCOUNT ' + str(cycles) + '\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        if sg_program == '29':
            # 1 full half cycle (inrush test), trip/No trip
            freq = float(sg_command[1])
            amplitude = self.sg_settings(sg_program, 1)
            self.sg_cmd = 'ARBLOAD ARB1\nDCOFFS 0\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nFREQ ' + str(freq) + '\nBSTPHASE 0\nBST NCYC\nBSTCOUNT 1\nSYNCTYPE BURST\nSYNCOUT ON\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        if sg_program == '31a':
            # DC signal, Trip/no trip
            # 10 Idn DC I
            amplitude = self.sg_settings(sg_program, 1)
            dc_offset = str(float(amplitude) - 0.01)
            self.sg_cmd = 'WAVE PULSE\nDCOFFS ' + dc_offset + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nPULSFREQ 9\nPULSSYMM 95\nPULSEDGE 0.00000001\nBST NCYC\nBSTCOUNT 1\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        if sg_program == '31b':
            # DC signal, Trip/no trip
            # 10 Idn DC II
            amplitude = self.sg_settings(sg_program, 1)
            dc_offset = str(float(amplitude) - 0.01)
            self.sg_cmd = 'WAVE PULSE\nDCOFFS -' + dc_offset + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nPULSFREQ 9\nPULSSYMM 95\nPULSEDGE 0.00000001\nBST NCYC\nBSTCOUNT 1\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT INVERT\nOUTPUT ON\n*TRG\n'
        if sg_program == '32a':
            # DC I ramp Trip Level
            dc_offset = self.sg_settings(sg_program, 1)
            dur_period = 1
            self.sg_cmd = 'WAVE PULSE\nAMPUNIT VRMS\nLOLVL 0\nHILVL 0.03\nDCOFFS -' + dc_offset + '\nPULSPER ' + str(dur_period) + '\nPULSSYMM 99\nPULSEDGE 0.00000001\nBST NCYC\nBSTCOUNT 1\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT INVERT\nOUTPUT ON\n*TRG\n'
        if sg_program == '	':
            # DC II ramp Trip Level
            HILVL = self.sg_settings(sg_program, 1)
            dur_period = 1
            self.sg_cmd = 'WAVE PULSE\nAMPUNIT VRMS\nLOLVL 0\nHILVL ' + HILVL + '\nPULSPER ' + str(dur_period) + '\nPULSSYMM 99\nPULSEDGE 0.00000001\nBST NCYC\nBSTCOUNT 1\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT ON\n*TRG\n'
        if sg_program == "34":  # AC 135 wave, Trip Level
            # AC 135 wave, Trip Time
            # Pulsating DC 135
            freq = float(sg_command[1])
            duration = float(sg_command[2]) / 1000  # in seconds
            amplitude = self.sg_settings(sg_program, 1)
            one_period = float(1 / freq)
            cycles = round(duration / one_period)
            self.sg_cmd = 'ARBLOAD ARB2\nDCOFFS 0\nFREQ ' + str(freq) + '\nAMPUNIT VRMS\nAMPL ' + str(amplitude) + '\nBST NCYC\nBSTCOUNT ' + str(cycles) + '\nSYNCOUT ON\nSYNCTYPE BURST\nTRGSRC MAN\nOUTPUT NORMAL\nOUTPUT ON\n*TRG\n'
        self.freq = str(freq)
        self.duration = str(duration)
        self.sg_setup_num = sg_program
        self.sp1 = str(p1)
        self.sp2 = str(p2)
        self.sp3 = str(p3)
        self.sp4 = str(p4)
        self.sp5 = str(p5)
        print("SG commmand = " + str(self.sg_cmd))
        self.sg_send(self.sg_cmd, False)
        # self.sg_send_optimized(self.sg_cmd)

    def sg_calibration(self, parameter):
        for calibration_line in self.product_calibration_lines:
            if int(calibration_line[0]) == int(self.test_number):
                return calibration_line[parameter]
        return -1

    def sg_settings(self, sg_program, parameter):
        for product_calibration_line in self.product_calibration_lines:
            if product_calibration_line[1] == sg_program:
                parameters = str(product_calibration_line[2]).split(',')
                return str(parameters[parameter - 1])
        return -1

    def exec_rsg_commmand(self):
        self.sg_cmd = 'SYNCOUT OFF\nOUTPUT OFF\nZLOAD OPEN\nOUTPUT NORMAL\nBST OFF\nAMPUNIT VRMS\nHILVL 10\nLOLVL 0\n'
        self.sg_send(self.sg_cmd, False)

    def sg_send_cmd_optimized(self, commands):
        new_commands = commands.replace("\n", ";")
        new_commands_list = commands.split(';')
        is_trg_in_new_commands_list = False
        if "*TRG" in new_commands_list:
            is_trg_in_new_commands_list = True
            new_commands_list.remove("*TRG")
        optimized_commands = ""
        for command in new_commands_list:
            if command in self.sg_commands:
                optimized_commands += (command + ";")
        if is_trg_in_new_commands_list:
            optimized_commands += "*TRG\n"
        # print("new parameters filtered:")
        print("Optimized SG commmand = " + optimized_commands)
        self.sg_send(commands, False)

    def sg_send(self, command, is_need_return):
        if self.sg_connection == "ser":
            self.sg_ser.writeData(bytearray(command.encode("UTF-8")))
            if is_need_return:
                result = "no_data"
                for wait_for_response in range(1, 10):
                    try:
                        if self.sg_ser.waitForReadyRead(200):
                            bs = bytes(self.sg_ser.readAll())
                            for b in bs:
                                result += chr(b)
                                if b == 10:
                                    return result
                    except Exception as e:
                        sleep(0.1)
                return result
        if self.sg_connection == "tcp":
            SocketMain(command, self.sg_ip_adr, "Signal")

    def mm_read(self, command, is_need_return):
        result = "no_data"
        if self.mm_connection == "ser":
            self.mm_ser.writeData(bytearray(command.encode("UTF-8")))
            if is_need_return:
                for wait_for_response in range(1, 10):
                    try:
                        if self.mm_ser.waitForReadyRead(200):
                            bs = bytes(self.mm_ser.readAll())
                            for b in bs:
                                result += chr(b)
                                if b == 10:
                                    result = result.replace("no_data", "")
                                    if result.find("OVLOAD") != -1:
                                        result = "OVLOAD"
                                    else:
                                        result = result[:result.find("e")]
                                    return result
                    except Exception as e:
                        sleep(0.1)
        if self.mm_connection == "tcp":
            result = SocketMain(str(command), self.mm_ip_adr, 'Multi')  # Setup new range
            return result

    def exec_mm_command(self, line):
        command = line[1]
        parameter_1 = parameter_2 = parameter_3 = parameter_4 = None
        if line[2] is not None:
            parameter_1 = line[2]
        if line[3] is not None:
            parameter_2 = line[3]
        if line[3] is not None:
            parameter_3 = line[3]
        if line[4] is not None:
            parameter_4 = line[4]
        self.pass_min = 0
        self.pass_max = 0
        new_mm_range = 0
        if command == "range":
            if parameter_1 is not None:
                self.pass_min = parameter_1
                self.pass_max = parameter_2
                self.mm_units = parameter_3
            else:
                self.pass_min = self.test_specification_lines[self.test_number-1][6]
                self.pass_max = self.test_specification_lines[self.test_number-1][7]
                self.mm_units = self.test_specification_lines[self.test_number-1][8]
            self.pMin = float(self.pass_min)
            self.pMax = float(self.pass_max)
            # self.mm_set_range(line)
            new_mm_range = 0
            cpd = ""
            self.mm_units = str(self.mm_units ).lower()
            if self.mm_units == "mvdc":
                if self.pMax < 100:
                    # print("DC Voltage measurement range 100mV ")
                    cpd = 'vdc 100mV\n'
                    new_mm_range = 1
                elif self.pMax < 1000:
                    # print("DC Voltage measurement range 1000mV ")
                    cpd = 'vdc 1000mV\n'
                    new_mm_range = 2

            if self.mm_units == "vdc":
                if self.pMax < 10:
                    # print("DC Voltage measurement range 10V ")
                    cpd = 'vdc 10V\n'
                    new_mm_range = 3
                elif self.pMax < 100:
                    # print("DC Voltage measurement range 100V ")
                    cpd = 'vdc 100V\n'
                    new_mm_range = 4
                elif self.pMax < 1000:
                    # print("DC Voltage measurement range 1000V ")
                    cpd = 'vdc 1000V\n'
                    new_mm_range = 5

            if self.mm_units == "mvac":
                if self.pMax < 100:
                    # print("AC Voltage measurement range 100mV ")
                    cpd = 'VAC 100mV\n'
                    new_mm_range = 6
                elif self.pMax < 1000:
                    # print("AC Voltage measurement range 1000mV ")
                    cpd = 'VAC 1000mV\n'
                    new_mm_range = 7

            if self.mm_units == "vac":
                if self.pMax < 10:
                    # print("AC Voltage measurement range 10V ")
                    cpd = 'VAC 10V\n'
                    new_mm_range = 8
                elif self.pMax < 100:
                    # print("AC Voltage measurement range 100V ")
                    cpd = 'VAC 100V\n'
                    new_mm_range = 9
                elif self.pMax < 1000:
                    # print("AC Voltage measurement range 1000V ")
                    cpd = 'VAC 1000V\n'
                    new_mm_range = 10

            if self.mm_units == "ma":
                if self.pMax < 1:
                    # print("Current measurement range 1mA ")
                    cpd = 'IAC 1mA\n'
                    new_mm_range = 11
                elif self.pMax < 100:
                    # print("Current measurement range 100mA ")
                    cpd = 'IAC 100mA\n'
                    new_mm_range = 12
                elif self.pMax < 1000:
                    # print("DC Current measurement range 1000mA ")
                    cpd = 'IAC 1000mA\n'
                    new_mm_range = 13

            if self.mm_units == "madc":
                if self.pMax < 1:
                    # print("DC Current measurement range 1mA ")
                    cpd = 'IDC 1mA\n'
                    new_mm_range = 14
                elif self.pMax < 100:
                    # print("DC Current measurement range 100mA ")
                    cpd = 'IDC 100mA\n'
                    new_mm_range = 15
                elif self.pMax < 1000:
                    # print("DC Current measurement range 1000mA ")
                    cpd = 'IDC 1000mA\n'
                    new_mm_range = 16

            if self.mm_units == "a":
                # print("Current measurement range 10 A ")
                cpd = 'IAC 10A\n'
                new_mm_range = 17

            if self.mm_units == "adc":
                # print("DC Current measurement range 10A ")
                cpd = 'IDC 10A\n'
                new_mm_range = 18

            if new_mm_range != self.mm_last_range:
                # Remember this range so next time will not be used if same range if requested
                self.mm_last_range = new_mm_range
                # Due to a bug in MM firmware after changing range, read command should be executed  :(
                cpd = cpd.replace("\n", ";read\n")
                measure = self.mm_read(cpd, True)   # Dummy read
                sleep(0.5)

        if command.find("read") != -1:
            if command == "read_emul_t":
                # Emulate multimetar pass value
                measure = (self.pMax + self.pMin) / 2
            elif command == "read_emul_f":
                # Emulate multimetar fail value
                measure = (self.pMax + self.pMin) / 3
            else:
                # Execute real measurement !!!!
                cpd = "read?\n"
                measure = self.mm_read(cpd, True)
                print("MM return:", measure)
            if command == "readmem":
                # Now just remember this value to be used for read- or read+ calculation
                self.last_value = float(measure)
            else:
                self.mm_value = measure
                if measure != "OVLOAD":
                    value = round(float(measure), 2)
                    if command == "read-":
                        value = round(self.last_value - value, 2)
                        self.last_value = round(float(measure), 2)
                        self.mm_value = str(value)
                    # Compare the result with pass criteria
                    if self.pMax > value > self.pMin:
                        self.is_mm_result = True
                        self.is_mm_error = False
                    else:
                        self.is_mm_result = False
                        self.is_mm_error = True
                else:
                    self.is_mm_result = False
                    self.is_mm_error = True
                self.log_mm_result()

    def mm_params(self, stest, parameter):
        for line in self.product_config_lines:
            test = line[:line.find("=")].replace(" ", "")
            if test == "m" + str(stest):
                parameters = line[line.find("=") + 1:].split(',')
                return str(parameters[parameter - 1])
        return -1

    def exec_trip_command(self, trip_command):
        # remove leading "strip" command
        del trip_command[0]
        board = trip_command[0]
        # remove leading "board name" parametar
        del trip_command[0]
        # remove trailing 2 parameters
        del trip_command[2]
        del trip_command[2]
        board_adr = 0
        if board == "ndb_1":
            board_adr = ndb_1
            self.slot = 1
        if board == "ndb_2":
            board_adr = ndb_2
            self.slot = 2
        if board == "ndb_3":
            board_adr = ndb_3
            self.slot = 3
        if board == "ndb_4":
            board_adr = ndb_4
            self.slot = 4

        if trip_command[0] is None:
            trip_command[0] = self.test_specification_lines[self.test_number - 1][6]
            trip_command[1] = self.test_specification_lines[self.test_number - 1][7]

        self.TripTime = 0
        for err_count in range(1, 6):
            try:
                req = self.modbus_client.read_holding_registers(0, 1, unit=board_adr)
                assert (not req.isError())
                self.TripTime = req.registers[0] / 10
                # Check is seconf parameter is empty ?
                if trip_command[1] is None:
                    # Second parameters is empty, so first parameter should be 't" or 'nt'
                    self.TripType = trip_command[0]
                    self.TripType = self.TripType.lower()
                    if self.TripType == "t" or self.TripType == 'trip':
                        # print("Slot "+ str(self.slot) + " TRIP required")
                        if self.TripTime > 0:
                            self.is_trip_result = True
                            self.is_trip_error = False
                        else:
                            self.is_trip_result = False
                            self.is_trip_error = True
                    if self.TripType == "nt" or self.TripType == 'no trip':
                        # print("Slot " + str(self.slot) + " NO trip required")
                        if self.TripTime == 0:
                            self.is_trip_result = True
                            self.is_trip_error = False
                        else:
                            self.is_trip_result = False
                            self.is_trip_error = True
                else:
                    self.TripType = "range"
                    pMin = float(trip_command[0])
                    pMax = float(trip_command[1])
                    self.pass_min = str(pMin)
                    self.pass_max = str(pMax)
                    if pMin < self.TripTime < pMax:
                        self.is_trip_error = False
                        self.is_trip_result = True
                    else:
                        self.is_trip_error = True
                        self.is_trip_result = False
                self.log_trip_result()
                # print("Slot " + str(self.slot) + " trip time = ", str(self.TripTime))
                break
            except Exception as e:
                print(e)
                self.log_comm_error("TRIP time read from " + board + ", attempt " + str(err_count))
                self.TripTime = -1

    def exec_wait_command(self, line):
        wait_time = line[1]
        if wait_time is None:
            wait_time = self.sg_calibration(4)
            print("Test duration time ", end="")
        print("wait " + str(wait_time) + " ms")
        sleep(int(wait_time) / 1000)

    @staticmethod
    def exec_pc_command(line):
        message = line[3]
        exec(message)

    @staticmethod
    def exec_print_command(line):
        message = line[3]
        print(message)

    def exec_msg_command(self, line):
        message = line[3]
        if message == "clr":
            message = " "
        self.ui.statusbar.showMessage(message)
        self.ui.statusbar.repaint()

    @staticmethod
    def what_is_ndb_relay_number(relay_name):
        # NEST DUT board have names for relays
        coil = 0
        if relay_name == "pmr-1" or relay_name == "11":
            coil = 11
        if relay_name == "pmr-2" or relay_name == "12":
            coil = 12
        if relay_name == "pmr-rst" or relay_name == "20":
            coil = 20
        if relay_name == "gpo-1" or relay_name == "15":
            coil = 15
        if relay_name == "gpo-2" or relay_name == "16":
            coil = 16
        if relay_name == "test_trg" or relay_name == "10":
            coil = 10
        if relay_name == "host-int" or relay_name == "19":
            coil = 19
        if relay_name == "led-p" or relay_name == "17":
            coil = 17
        if relay_name == "led-f" or relay_name == "18":
            coil = 18
        if relay_name == "cd-1" or relay_name == "13":
            coil = 13
        if relay_name == "cd-2" or relay_name == "14":
            coil = 14
        if relay_name == "pwr" or relay_name == "9":
            coil = 9
        if relay_name == "tp-1" or relay_name == "1":
            coil = 1
        if relay_name == "tp-2" or relay_name == "2":
            coil = 2
        if relay_name == "tp-3" or relay_name == "3":
            coil = 3
        if relay_name == "tp-4" or relay_name == "4":
            coil = 4
        if relay_name == "tp-5" or relay_name == "5":
            coil = 5
        if relay_name == "tp-6" or relay_name == "6":
            coil = 6
        if relay_name == "tp-7" or relay_name == "7":
            coil = 7
        if relay_name == "tp-8" or relay_name == "8":
            coil = 8
        return coil

    # barcode, slot_ID, test num, pass_min, pass_max, mm_value, units, test_result
    def log_mm_result(self):
        barcode = ""
        # change color to red/green of test numbers for corrent slot (30 rectangles in row below bar code)
        pb_test = "pb_" + str(self.slot) + "_" + str(self.test_number)
        exec_str = 'self.set_btn_pass_fail_color("' + pb_test + '", self.is_mm_result)'
        exec(exec_str)
        exec("self.ui." + pb_test + ".repaint()")
        test_num = int(self.test_number)
        pf_string = "PASS"
        if not self.is_mm_result:
            pf_string = "FAIL"
            # change color to red of test numbers for slot/test
            pb_pf = "pb_pf_t" + str(self.test_number) + "_" + str(self.slot)
            exec('self.set_btn_pass_fail_color("' + pb_pf + '", False)')
            exec("self.ui." + pb_pf + ".repaint()")
            # Increase numbers of all errors for current test for all 4 slots
            self.fail_units_counters_per_test[test_num] += 1
            lbl_fail_t = "lbl_fail_t" + str(self.test_number)
            exec_str = "self.ui." + lbl_fail_t + ".setText(str(self.fail_units_counters_per_test[test_num]))"
            exec(exec_str)
            exec("self.ui." + lbl_fail_t + ".repaint()")
            if self.slot == 1:
                self.fail_units_counters_per_test_for_slot_1[test_num] += 1
                exec_str = "self.ui.pb_pf_t" + str(self.test_number) + "_1.setText('" + str(self.fail_units_counters_per_test_for_slot_1[test_num]) + "')"
                exec(exec_str)
            if self.slot == 2:
                self.fail_units_counters_per_test_for_slot_2[test_num] += 1
                exec_str = "self.ui.pb_pf_t" + str(self.test_number) + "_2.setText('" + str(self.fail_units_counters_per_test_for_slot_2[test_num]) + "')"
                exec(exec_str)
            if self.slot == 3:
                self.fail_units_counters_per_test_for_slot_3[test_num] += 1
                exec_str = "self.ui.pb_pf_t" + str(self.test_number) + "_3.setText('" + str(self.fail_units_counters_per_test_for_slot_3[test_num]) + "')"
                exec(exec_str)
            if self.slot == 4:
                self.fail_units_counters_per_test_for_slot_4[test_num] += 1
                exec_str = "self.ui.pb_pf_t" + str(self.test_number) + "_4.setText('" + str(self.fail_units_counters_per_test_for_slot_4[test_num]) + "')"
                exec(exec_str)
        if self.slot == 1:
            # Update 3 labels under barcode ("Last test", "Measurement", "Status") for slot 1
            barcode = self.barcode_slot_1
            self.ui.lbl_test_no_1.setText(str(self.test_number))
            self.ui.lbl_measurement_1.setText(str(self.mm_value) + " " + self.mm_units)
            self.ui.lbl_ResultTest_1.setText(pf_string)
            self.test_pass_result_slot_1[test_num-1] = self.is_mm_result
            self.test_value_result_slot_1[test_num-1] = self.mm_value
        if self.slot == 2:
            # Update 3 labels under barcode ("Last test", "Measurement", "Status") for slot 2
            barcode = self.barcode_slot_2
            self.ui.lbl_test_no_2.setText(str(self.test_number))
            self.ui.lbl_measurement_2.setText(str(self.mm_value) + " " + self.mm_units)
            self.ui.lbl_ResultTest_2.setText(pf_string)
            self.test_pass_result_slot_2[test_num-1] = self.is_mm_result
            self.test_value_result_slot_2[test_num-1] = self.mm_value
        if self.slot == 3:
            # Update 3 labels under barcode ("Last test", "Measurement", "Status") for slot 3
            barcode = self.barcode_slot_3
            self.ui.lbl_test_no_3.setText(str(self.test_number))
            self.ui.lbl_measurement_3.setText(str(self.mm_value) + " " + self.mm_units)
            self.ui.lbl_ResultTest_3.setText(pf_string)
            self.test_pass_result_slot_3[test_num-1] = self.is_mm_result
            self.test_value_result_slot_3[test_num-1] = self.mm_value
        if self.slot == 4:
            # Update 3 labels under barcode ("Last test", "Measurement", "Status") for slot 4
            barcode = self.barcode_slot_4
            self.ui.lbl_test_no_4.setText(str(self.test_number))
            self.ui.lbl_measurement_4.setText(str(self.mm_value) + " " + self.mm_units)
            self.ui.lbl_ResultTest_4.setText(pf_string)
            self.test_pass_result_slot_4[test_num-1] = self.is_mm_result
            # self.test_value_result_slot_4[test_num-1] = float(self.mm_value)
            self.test_value_result_slot_4[test_num-1] = self.mm_value
        barcode = barcode[:17]
        # Get time stump
        dt = str(datetime.now())
        dt = dt[:dt.find(".")]
        log_line = "mm," + \
                   dt + "," + \
                   barcode + "," + \
                   str(self.slot) + "," + \
                   str(self.test_number) + "," + \
                   str(self.pMin) + "," + \
                   str(self.pMax) + "," + \
                   str(self.mm_value) + "," + \
                   self.mm_units + "," + \
                   pf_string + "\n"
        self.log_lines.append(log_line)
        self.mm_log_lines.append(log_line)

    # barcode, slot_ID, test num, freq, fault_current, pass_min, pass_max, trip_time, trip_type, sg_setup_num, sp1, sp2, sp3, sp4, sp5, sp6, sp7
    # test_result
    def log_trip_result(self):
        tts = str(self.TripTime) + "ms"
        pb_test = "pb_" + str(self.slot) + "_" + str(self.test_number)
        exec_str = 'self.set_btn_pass_fail_color("' + pb_test + '", self.is_trip_result)'
        exec(exec_str)
        exec("self.ui." + pb_test + ".repaint()")
        test_num = int(self.test_number)
        pf_string = "PASS"
        if not self.is_trip_result:
            pf_string = "FAIL"
            pb_pf = "pb_pf_t" + str(self.test_number) + "_" + str(self.slot)
            exec('self.set_btn_pass_fail_color("' + pb_pf + '", False)')
            exec("self.ui." + pb_pf + ".repaint()")
            self.fail_units_counters_per_test[test_num] += 1
            lbl_fail_t = "lbl_fail_t" + str(self.test_number)
            exec_str = "self.ui." + lbl_fail_t + ".setText(str(self.fail_units_counters_per_test[test_num]))"
            exec(exec_str)
            exec("self.ui." + lbl_fail_t + ".repaint()")
            if self.slot == 1:
                self.fail_units_counters_per_test_for_slot_1[test_num] += 1
                exec_str = "self.ui.pb_pf_t" + str(self.test_number) + "_1.setText('" + str(self.fail_units_counters_per_test_for_slot_1[test_num]) + "')"
                exec(exec_str)
            if self.slot == 2:
                self.fail_units_counters_per_test_for_slot_2[test_num] += 1
                exec_str = "self.ui.pb_pf_t" + str(self.test_number) + "_2.setText('" + str(self.fail_units_counters_per_test_for_slot_2[test_num]) + "')"
                exec(exec_str)
            if self.slot == 3:
                self.fail_units_counters_per_test_for_slot_3[test_num] += 1
                exec_str = "self.ui.pb_pf_t" + str(self.test_number) + "_3.setText('" + str(self.fail_units_counters_per_test_for_slot_3[test_num]) + "')"
                exec(exec_str)
            if self.slot == 4:
                self.fail_units_counters_per_test_for_slot_4[test_num] += 1
                exec_str = "self.ui.pb_pf_t" + str(self.test_number) + "_4.setText('" + str(self.fail_units_counters_per_test_for_slot_4[test_num]) + "')"
                exec(exec_str)
        barcode = ""
        if self.slot == 1:
            barcode = self.barcode_slot_1
            self.ui.lbl_test_no_1.setText(str(self.test_number))
            self.ui.lbl_measurement_1.setText(tts + " " + self.TripType)
            self.ui.lbl_ResultTest_1.setText(pf_string)
            self.test_pass_result_slot_1[test_num-1] = self.is_trip_result
            self.test_value_result_slot_1[test_num-1] = self.TripTime
        if self.slot == 2:
            barcode = self.barcode_slot_2
            self.ui.lbl_test_no_2.setText(str(self.test_number))
            self.ui.lbl_measurement_2.setText(tts + " " + self.TripType)
            self.ui.lbl_ResultTest_2.setText(pf_string)
            self.test_pass_result_slot_2[test_num-1] = self.is_trip_result
            self.test_value_result_slot_2[test_num-1] = self.TripTime
        if self.slot == 3:
            barcode = self.barcode_slot_3
            self.ui.lbl_test_no_3.setText(str(self.test_number))
            self.ui.lbl_measurement_3.setText(tts + " " + self.TripType)
            self.ui.lbl_ResultTest_3.setText(pf_string)
            self.test_pass_result_slot_3[test_num-1] = self.is_trip_result
            self.test_value_result_slot_3[test_num-1] = self.TripTime
        if self.slot == 4:
            barcode = self.barcode_slot_4
            self.ui.lbl_test_no_4.setText(str(self.test_number))
            self.ui.lbl_measurement_4.setText(tts + " " + self.TripType)
            self.ui.lbl_ResultTest_4.setText(pf_string)
            self.test_pass_result_slot_4[test_num-1] = self.is_trip_result
            self.test_value_result_slot_4[test_num-1] = self.TripTime
        barcode = barcode[:17]
        barcode = barcode.strip(chr(0))
        # Get time stump
        dt = str(datetime.now())
        dt = dt[:dt.find(".")]
        log_line = "tt," + \
                   dt + "," + \
                   barcode + "," + \
                   str(self.slot) + "," + \
                   str(self.test_number) + "," + \
                   str(self.freq) + "," + \
                   self.fc_setup + "," + \
                   self.pass_min + "," + \
                   self.pass_max + "," + \
                   str(self.TripTime) + "," + \
                   self.TripType + "," + \
                   self.duration + "," + \
                   self.sg_setup_num + "," + \
                   self.sp1 + "," + \
                   self.sp2 + "," + \
                   self.sp3 + "," + \
                   self.sp4 + "," + \
                   self.sp5 + "," + \
                   pf_string + "\n"
        self.log_lines.append(log_line)
        self.tt_log_lines.append(log_line)

    def save_log_file(self):
        file = open(self.debug_log_file_name, "a+")
        file.writelines(self.log_lines)
        file.close()
        self.log_lines.clear()

    @staticmethod
    def set_bit(value, bit):
        return value | (1 << bit)

    @staticmethod
    def clear_bit(value, bit):
        return value & ~(1 << bit)

    def save_new_format_log_file(self):
        # print("Start new format log file")

        # Date, Time, Station, Product, Batch, IDUser, Position, Board, Test1,.., Test30, Result
        file = open(self.log_file_name_new_format, "a+")

        dt = str(datetime.now())
        dt = dt[:dt.find(".")]
        log_line_header = dt + "," \
                          + self.tester_id + "," \
                          + self.user_id + ","  \
                          + self.product_name + "," \
                          + self.nest_id + "," \
                          + self.insert_id + "," \
                          + self.grn_no + "," \
                          + self.test_specification_file_issue + "," \
                          + self.test_specification_file_ref + "," \
                          + self.test_sequence_file_issue + "," \
                          + self.test_sequence_file_ref + "," \
                          + self.test_calibration_file_ref + "," \
                          + self.app_ver + "," \
                          + ","
        # Slot 1 log line
        log_line = log_line_header + "1,r"
        log_line += self.barcode_slot_1.strip(chr(0))
        for test_value in self.test_value_result_slot_1:
            log_line += "," + str(test_value)
        if self.is_slot_1_test_pass:
            log_line += ",0"
        else:
            results_str_semicolumn_delimited = ""
            for test_num in range(0, 29):
                if not self.test_pass_result_slot_1[test_num]:
                    results_str_semicolumn_delimited += str(test_num + 1) + ";"
            log_line += "," + str(results_str_semicolumn_delimited)
            last_semicolumn_pos = log_line.rfind(";")
            log_line = log_line[:last_semicolumn_pos] + log_line[last_semicolumn_pos+1:]
        log_line += "\n"
        file.write(log_line)

        if self.insert_number_duts > 1:
            # Slot 2 log line
            log_line = log_line_header + "2,r"
            log_line += self.barcode_slot_2.strip(chr(0))
            for test_value in self.test_value_result_slot_2:
                log_line += "," + str(test_value)
            if self.is_slot_2_test_pass:
                log_line += ",0"
            else:
                results_str_semicolumn_delimited = ""
                for test_num in range(0, 29):
                    if not self.test_pass_result_slot_2[test_num]:
                        results_str_semicolumn_delimited += str(test_num + 1) + ";"
                log_line += "," + results_str_semicolumn_delimited
                last_semicolumn_pos = log_line.rfind(";")
                log_line = log_line[:last_semicolumn_pos] + log_line[last_semicolumn_pos + 1:]
            log_line += "\n"
            file.write(log_line)
        
        if self.insert_number_duts > 2:
            # Slot 3 log line
            log_line = log_line_header + "3,r"
            log_line += self.barcode_slot_3.strip(chr(0))
            for test_value in self.test_value_result_slot_3:
                log_line += "," + str(test_value)
            if self.is_slot_3_test_pass:
                log_line += ",0"
            else:
                results_str_semicolumn_delimited = ""
                for test_num in range(0, 29):
                    if not self.test_pass_result_slot_3[test_num]:
                        results_str_semicolumn_delimited += str(test_num + 1) + ";"
                log_line += "," + results_str_semicolumn_delimited
                last_semicolumn_pos = log_line.rfind(";")
                log_line = log_line[:last_semicolumn_pos] + log_line[last_semicolumn_pos + 1:]
            log_line += "\n"
            file.write(log_line)
        
        if self.insert_number_duts > 3:
            # Slot 4 log line
            log_line = log_line_header + "4,r"
            log_line += self.barcode_slot_4.strip(chr(0))
            for test_value in self.test_value_result_slot_4:
                log_line += "," + str(test_value)
            if self.is_slot_4_test_pass:
                log_line += ",0"
            else:
                results_str_semicolumn_delimited = ""
                for test_num in range(0, 29):
                    if not self.test_pass_result_slot_4[test_num]:
                        results_str_semicolumn_delimited += str(test_num+1) + ";"
                log_line += "," + results_str_semicolumn_delimited
                last_semicolumn_pos = log_line.rfind(";")
                log_line = log_line[:last_semicolumn_pos] + log_line[last_semicolumn_pos + 1:]
            log_line += "\n"
            file.write(log_line)
        file.close()
        # print("end new format log file")

    def log_comm_error(self, board_error):
        file = open(self.comm_error_log_file_name, "a+")
        dt = str(datetime.now())
        dt = dt[:dt.find(".")]
        board_error = dt + " " + board_error
        file.write(board_error + "\n")
        file.close()

    def log_bar_code_error(self, slot, attempt):
        file = open(self.bc_error_log_file_name, "a+")
        dt = str(datetime.now())
        dt = dt[:dt.find(".")]
        bar_code_error = dt + "," + str(slot) + "," + str(attempt)
        file.write(bar_code_error + "\n")
        file.close()

    def log_mm_comm_error(self, event):
        file = open(self.mm_error_log_file_name, "a+")
        dt = str(datetime.now())
        dt = dt[:dt.find(".")]
        bar_code_error = dt + "," + event
        file.write(bar_code_error + "\n")
        file.close()


"""
    Main application execute part .... 
"""

if __name__ == "__main__":
    app = QApplication(sys.argv)
    myapp = MainAppGui()
    myapp.show()
    sys.exit(app.exec_())

"""
TIPS 
To run app remotely, execute this command in terminal if connected over SSH
DISPLAY=:0.0 python3 app.py 

http://www.gusnan.se/devilspie2/
http://www.gusnan.se/devilspie2/manual.php#set_window_position

Usefil links
Qt Style Sheets Examples
http://qt.developpez.com/doc/4.6/stylesheet-examples/#customizing-the-foreground-and-background-colors
https://doc.qt.io/qt-5/stylesheet-examples.html

CSS COLOR NAMES
https://www.w3schools.com/cssref/css_colors.asp
http://cng.seas.rochester.edu/CNG/docs/x11color.html

WIRING PI
https://github.com/WiringPi/WiringPi2-Python
http://wiringpi.com/
import wiringpi2

"""


"""
https://github.com/splitbrain/rpibplusleaf
https://stackoverflow.com/questions/48511765/detecting-mechanical-button-press-with-python3-and-qt5/48512467#48512467

"""

"""
# http://www.instructables.com/id/Raspberry-Pi-Launch-Python-script-on-startup

Running a Python Script at Boot
Pokretanje sa desktopom....
https://jackbarber.co.uk/blog/2017-03-02-automatically-run-a-python-script-at-boot-in-raspbian

Not succeded with etc/rc.local, but this have to be run also "sudo crontab -e", here s solution :
https://www.raspberrypi.org/forums/viewtopic.php?f=28&t=187868

https://www.raspberrypi.org/forums/viewtopic.php?f=31&t=43509
"""

# test=1,2,3,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25
# test=1,2,3,5,6,7,8,9,10,11,12,13,14,25


# WA Site Ground hosting
# testdata@westernautomation.com
# 0fKCBz32Um8F

# FTP
# tester_01
# Tester_01-password
# Home folder /home/watestda/public_html/tester_01

# MySql
# Domain:
# Primary Domain: watestdata.com
# Database: watestda_test
# User: watestda_test
# Password: Watestda_test-Password-8299

# port 3306